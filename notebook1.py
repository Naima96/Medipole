# -*- coding: utf-8 -*-
"""
Created on Fri Jan 15 10:51:20 2021

@author: al-abiad
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Jan 14 16:38:51 2021

@author: al-abiad
"""
import matplotlib.pyplot as plt
import os
from os import walk
from scipy import signal,fft
from scipy.signal import hilbert
import tkinter as tk
import pygubu
from tkinter import messagebox
from CPclass import phone as CP
from scipy.signal import find_peaks
import numpy as np

from matplotlib.backends.backend_tkagg import ( FigureCanvasTkAgg, NavigationToolbar2Tk)# Implement the default Matplotlib key bindings.
from matplotlib.figure import Figure

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

#nonlinear
from nolitsa import data, delay,dimension
import nolds as nld

CURRENT_DIR = os.path.abspath(os.path.dirname(__file__))
pickable_artists1 = []
pickable_artists2 = []
pickable_artists3=[]

class MyApplication:
    
    def __init__(self):
        
        self.walking_period=[]

        #1: Create a builder
        self.builder = builder = pygubu.Builder()

        #2: Load an ui file
        builder.add_from_file(os.path.join(CURRENT_DIR, 'notebook23.ui'))
        
        #3: Create the toplevel widget.
        self.mainwindow = builder.get_object('mainwindow')

        # Container for the matplotlib canvas and toolbar classes
        self.fcontainer = builder.get_object('fcontainer')
        self.fcontainer2 = builder.get_object('fcontainer2')
        self.fcontainer3 = builder.get_object('fcontainer3')
        self.fcontainer4 = builder.get_object('fcontainer4')
        self.fcontainer_gaitupfeet= builder.get_object('fcontainer_gaitupfeet')
        self.fcontainer_gaituphand= builder.get_object('fcontainer_gaituphand')
        self.fcontainer_phonewaist= builder.get_object('fcontainer_phonewaist')
        self.fcontainer_allsignals=builder.get_object('fcontainer_allsignals')
        self.fcontainer_signalwithstep=builder.get_object('fcontainer_signalwithstep')
        #set filepathchooser
        self.filepath = builder.get_object('filepath')
        self.filepath_export=builder.get_object('path_choose_Export')
        
        #get labels
        self.lbl_files=builder.get_object('lbl_files')
        self.lbl_SDstride=builder.get_object('SD_stride')
        self.lbl_CVstride=builder.get_object('CV_stride')
        self.lbl_Nstride=builder.get_object('N_stride')
        self.lbl_Nphase=builder.get_object('N_phase')
        self.lbl_mean_stridetime=builder.get_object('lbl_mean_stridetime')
        
        self.lbl_SDstride_GU=builder.get_object('SD_stride_GU')
        self.lbl_CVstride_GU=builder.get_object('CV_stride_GU')
        self.lbl_Nstride_GU=builder.get_object('N_stride_GU')
        self.lbl_Nphase_GU=builder.get_object('N_phase_GU')
        self.lbl_mean_stridetime_GU=builder.get_object('lbl_mean_stridetime_GU')
        
        self.lbl_SDstride_HP=builder.get_object('SD_stride_HP')
        self.lbl_CVstride_HP=builder.get_object('CV_stride_HP')
        self.lbl_Nstride_HP=builder.get_object('N_stride_HP')
        self.lbl_Nphase_HP=builder.get_object('N_phase_HP')
        self.lbl_mean_stridetime_HP=builder.get_object('lbl_mean_stridetime_HP')
        
        self.lbl_lyapx=builder.get_object('lbl_lyapx')
        self.lbl_lyapy=builder.get_object('lbl_lyapy')
        self.lbl_lyapz=builder.get_object('lbl_lyapz')
        self.lbl_lyapr=builder.get_object('lbl_lyapr')
        
        self.lbl_se_x=builder.get_object('se_x')
        self.lbl_se_y=builder.get_object('se_y')
        self.lbl_se_z=builder.get_object('se_z')
        self.lbl_se_r=builder.get_object('se_r')
        
        self.lbl_se_stridetime=builder.get_object('se_stridetime')
        self.lbl_se_gc_x=builder.get_object('se_gc_x')
        self.lbl_se_gc_y=builder.get_object('se_gc_y')
        self.lbl_se_gc_z=builder.get_object('se_gc_z')
        self.lbl_se_gc_r=builder.get_object('se_gc_r')
        
        
        self.lbl_DFA_phone=builder.get_object('lbl_dfaphone')
        self.lbl_DFA_feet=builder.get_object('lbl_dfafeet')
        self.lbl_hand=builder.get_object('lbl_dfahand')

        #get entries
        self.ent_str_walk = self.builder.get_object('str_walk')
        self.ent_stp_walk = self.builder.get_object('stp_walk')
        
        self.ent_turn = self.builder.get_object('step_turn')
        self.ent_embdim = self.builder.get_object('ent_embdim')
        self.ent_timedelay = self.builder.get_object('ent_timedelay')
        self.ent_Nbstridelyap = self.builder.get_object('ent_Nbstridelyap')
        self.ent_turntime=self.builder.get_object('ent_turntime')
        
        self.ent_vect_length = self.builder.get_object('vect_length')
        self.ent_tolerance=self.builder.get_object('tol_r')
        
        
        
        self.ent_str_gaitup_feet=self.builder.get_object('str_gaitup_feet')
        self.ent_str_phone_waist=self.builder.get_object('str_phone_waist')
        self.ent_str_gaitup_hand=self.builder.get_object('str_gaitup_hand')
        
        self.ent_stp_gaitup_feet=self.builder.get_object('end_foot')
        self.ent_stp_phone_waist=self.builder.get_object('end_phone')
        self.ent_stp_phone_gaitup=self.builder.get_object('end_both')
        
        
        #get treeview
        self.treeview=self.builder.get_object('myetv')
        self.treeview_res=self.builder.get_object('Result_treeview')
        self.treeview_stat=self.builder.get_object('stat_treeview')
        
        self.treeview_stat_GU=self.builder.get_object('stat_treeview_GU')
        self.treeview_stat_HP=self.builder.get_object('stat_treeview_HP')
        
        
        #buttons
        # self.btn_cal_norm=self.builder.get_object('btn_cal_norm')
        # self.btn_plot=self.builder.get_object('btn_plot')
        self.btn_detect_start_stop=self.builder.get_object('btn_detect_start_stop')
        self.btn_detect_step=self.builder.get_object('btn_detect_step')
        self.btn_detect_turn=self.builder.get_object('btn_detect_turn')
        self.btn_cal_lyap=self.builder.get_object('btn_cal_lyap')
        self.btn_cal_time_delay=self.builder.get_object('btn_cal_time_delay')
        self.btn_cal_emb_dim=self.builder.get_object("btn_cal_emb_dim")
        
        #radio
        self.var_turn= self.builder.get_variable('var_turn')
        self.var_stepdetectionmethod= self.builder.get_variable('var_step_method')

        

        # Connect button callback
        builder.connect_callbacks(self)
        
        


        
    def on_click_canvas(self,event,mode="detect_turns"):
        
        if self.left_excel or self.right_excel:
            mode="detect_turns"
        else:
            mode="detect_peaks"
            
        

#        print('-----')
#        print('button:', event.button)
#        print('xdata, ydata:', event.xdata, event.ydata)
#        print('x, y:', event.x, event.y)
#        print('canvas:', event.canvas)
        print('inaxes:', event.inaxes)
        
        if self.toolbar.mode =="":
            print("toolbar is not selected")
            if mode=="detect_walking_period":
                print("detecting walking period")
                # if event.inaxes is not None and not hasattr(event, 'already_picked'):
                #     remove1 = [artist for artist in pickable_artists1 if artist.contains(event)[0]]
                #     print("--remove--")
                #     print(remove1)
                #     print("---pickable_artists---")
                #     print(pickable_artists1)
                #     x=event.xdata.astype('int')
                #     if not remove1:
                #         # add a pt
                #         pt1 = self.a.scatter(x, 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
                #         pickable_artists1.append(pt1)
                #         self.walking_period.append(x)
                #         print(self.walking_period)
                #     else:
                #         #remove a point
                #         pickable_artists.remove(remove[0])
                #         self.walking_period=[i for i in self.walking_period if i>x+50 or i<x-50]
                #         for artist in remove:
                #             artist.remove()
                #     self.canvas.draw()
                # else:
                #     print ('Clicked ouside axes bounds but inside plot window')
            if mode=="detect_peaks":
                print("detecting walking period")
                if event.inaxes is not None and not hasattr(event, 'already_picked'):
                    remove3 = [artist for artist in pickable_artists3 if artist.contains(event)[0]]
                    print("--remove1--")
                    print(remove3)
                    print("---pickable_artists---")
                    print(pickable_artists3)
                    x=event.xdata.astype('int')
                    
                    if not remove3:
                        # add a pt
                        pt3 = self.a.scatter(x, 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
                        pickable_artists3.append(pt3)
                        self.turn_peak_ind=np.append(self.turn_peak_ind,x)
                        # self.refresh_treeview()
                    else:
                        self.turn_peak_ind=[i for i in self.turn_peak_ind if i>x+100 or i<x-100]
                        # self.refresh_treeview()
                        pickable_artists3.remove(remove3[0])
                        for artist in remove3:
                            artist.remove()
                    self.canvas.draw()
                else:
                    print ('Clicked ouside axes bounds but inside plot window')
                
            
            elif mode=="detect_turns":
                print("detect_turns")
                print(self.var_turn.get())
                if self.var_turn.get()=="Stop_walk":
                    print("Stop_walk")
                    if event.inaxes is not None and not hasattr(event, 'already_picked'):
                        remove1 = [artist for artist in pickable_artists1 if artist.contains(event)[0]]
                        print("--remove1--")
                        print(remove1)
                        print("---pickable_artists---")
                        print(pickable_artists1)
                        x=event.xdata.astype('int')
                        
                        if not remove1:
                            # add a pt
                            pt1 = self.a.scatter(x, 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
                            pickable_artists1.append(pt1)
                            self.turn_start_ind=np.append(self.turn_start_ind,x)
                            # self.refresh_treeview()
                        else:
                            self.turn_start_ind=[i for i in self.turn_start_ind if i>x+100 or i<x-100]
                            # self.refresh_treeview()
                            pickable_artists1.remove(remove1[0])
                            for artist in remove1:
                                artist.remove()
                        self.canvas.draw()
                    else:
                        print ('Clicked ouside axes bounds but inside plot window')
                    
                elif self.var_turn.get()=="Start_walk":
                    print("Start_walk")
                    
                    if event.inaxes is not None and not hasattr(event, 'already_picked'):
                        remove2 = [artist for artist in pickable_artists2 if artist.contains(event)[0]]
                        print("--remove2--")
                        print(remove2)
                        print("---pickable_artists---")
                        print(pickable_artists2)
                        x=event.xdata.astype('int')
                        
                        if not remove2:
                            # add a pt
                            pt2 = self.a.scatter(x, 10,color='g',marker="|",s=10000000000,zorder=2,picker=5)
                            pickable_artists2.append(pt2)
                            self.turn_stop_ind=np.append(self.turn_stop_ind,x)
                            # self.refresh_treeview()
                        else:
                            self.turn_stop_ind=[i for i in self.turn_stop_ind if i>x+100 or i<x-100]
                            # self.refresh_treeview()
                            pickable_artists2.remove(remove2[0])
                            for artist in remove2:
                                artist.remove()
                        self.canvas.draw()
                    else:
                        print ('Clicked ouside axes bounds but inside plot window')

        else:
            print("toolbar is selected")
            
            
    def Refresh_treeview(self,event=None):
        print("refreshing...")
        if self.left_excel or self.right_excel:
            mode="detect_turns"
        else:
            mode="detect_peaks"
        
        if mode=="detect_turns":
            turn=int(self.ent_turn.get())
            stop_turn=np.sort(self.turn_stop_ind)
            start_turn=np.sort(self.turn_start_ind)
            
            if len(self.turn_stop_ind)!=len(self.turn_stop_ind):
                print("maybe you will face a problem in table")
                length=np.minimum(len(start_turn),len(stop_turn))
                
            length=np.minimum(len(start_turn),len(stop_turn))   
            self.phase=[]
               
            
            for i in range(0,length-1):
                self.phase.append((i+1,stop_turn[i],start_turn[i],0))
                
        elif mode=="detect_peaks":
            turn=int(self.ent_turn.get())
            self.turn_peak_ind.sort()
            self.phase=[]
            length=len(self.turn_peak_ind)  
            for i in range(0,length-1):
                self.phase.append((i+1,self.turn_peak_ind[i],self.turn_peak_ind[i+1],turn))

            
        if (len(self.treeview.get_children())!=0):
            self.treeview.delete(*self.treeview.get_children())
           
        for d in self.phase:
            self.treeview.insert('', tk.END, values=d)
       
        
    def on_path_changed(self, event=None):
        # Get the path choosed by the user
        print("you are in on path changed")
        self.lyap_NBstrides=0
        self.preprocessed=False
        self.btn_detect_start_stop["state"] = "disabled"
        self.btn_detect_step["state"] = "disabled"
        self.btn_detect_turn["state"] = "disabled"
        
        self.path = self.filepath.cget('path')
        self.export_path=self.path
        
        phone_dir=os.path.join(self.path,"telephone" )
        for root, dirs, files in os.walk(phone_dir):
            for file in files:
                if file.endswith(".txt"):
                    self.phone_direct=root
                break
        #phone directory is self.phone_direct
        print(self.phone_direct)
        
        #gaitup directory
        gaitup_dir=os.path.join(self.path,"Gaitup" )
        
        # show paths in label
        _, _, self.files = next(walk(self.phone_direct))
        self.lbl_files.configure(text = 'The files are %s ,%s and %s'%(self.files[0],self.files[1],self.files[2]))
        
        
    def on_path_changed_export(self, event=None):
        
        self.export_path= self.filepath_export.cget('path')
        print("export path is changed")

    
    def Detect_startandend(self,event=None):
        print("nothing")
        var_turn= self.builder.get_variable('var_turn')
        print(var_turn.get())

    #     walk_end=int(len(self.CP_data.acc_magnitude)-self.CP_data.detectstartofwalk(sig1=self.CP_data.acc_magnitude[::-1])*100)//100
    #     print(walk_end)
    #     walk_start=self.CP_data.detectstartofwalk(sig1=self.CP_data.acc_magnitude)
    #     print(walk_start)
    #     self.ent_str_walk.delete(0,tk.END)
    #     self.ent_str_walk.insert(0,walk_start)
    #     self.ent_stp_walk.delete(0,tk.END)
    #     self.ent_stp_walk.insert(0,walk_end)
    #     self.btn_detect_turn["state"] = "normal"
        
    def Detect_turns(self,method="gyro_peak",event=None):
        if self.left_excel or self.right_excel:
            method="extract gaitup"
        else:
            method=="gyro_peak"
        if method=="gyro_peak":
            self.phase=[]
            str_walk=0
            
            s_w=int((self.ent_str_phone_waist.get()))
            
            stop_phone=int((self.ent_stp_walk.get()))-s_w
            stp_walk=stop_phone
            
            turn=int(self.ent_turn.get())
            
            self.CP_data_phone.peakdet_m2(acc=False,plot_peak=True,detect_turn=True)
            self.peaks=self.CP_data_phone.peakandvalley["peak_index"]
    
            self.peaks=self.peaks[self.peaks>str_walk]
            self.peaks=self.peaks[self.peaks<stp_walk]
            
            self.peaks2 = [str_walk, *self.peaks, stp_walk]
    
            for i in range(0,len(self.peaks2)-1):
                self.phase.append((i+1,self.peaks2[i],self.peaks2[i+1],turn))
             
            self.peaks2=np.array(self.peaks2)
            
            if (len(self.treeview.get_children())!=0):
                self.treeview.delete(*self.treeview.get_children())
                
            for d in self.phase:
                self.treeview.insert('', tk.END, values=d)
    
            self.btn_detect_step["state"] = "normal"
            
            print(len(self.treeview.get_children()))
            
            for p in self.phase:
                pt = self.a.scatter(p[1], 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
                pickable_artists3.append(pt)
                
                pt = self.a.scatter(p[2], 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
                pickable_artists3.append(pt)
            
            self.turn_peak_ind=self.peaks2
            
            self.canvas.draw()
        elif method=="extract gaitup":
            print("extracting from gaitup")
            start_turn=self.turning_periods_left_right[:,0]
            stop_turn=self.turning_periods_left_right[:,1]
            self.phase=[]
            
            self.phase.append((0,0,start_turn[0],0))
            
            for i in range(0,len(start_turn)-1):
                if stop_turn[i]==start_turn[i]:
                    self.phase.append((i+1,stop_turn[i]+100,start_turn[i+1],0))
                else:   
                    self.phase.append((i+1,stop_turn[i],start_turn[i+1],0))
                     
            self.phase.append((i+2,stop_turn[-1],len(self.CP_data_phone.acc_magnitude),0))
            
            if (len(self.treeview.get_children())!=0):
                self.treeview.delete(*self.treeview.get_children())
                
            for d in self.phase:
                self.treeview.insert('', tk.END, values=d)
                
            for p in self.phase:
                print("in plot")
                # self.a.axvspan(p[1],p[2],color='red',alpha=0.25)
                # self.b.axvspan(p[1],p[2],color='red',alpha=0.25)
                
                #stop turn
                pt2 = self.a.scatter(p[1], 10,color='g',marker="|",s=10000000000,zorder=2,picker=5)
                pickable_artists2.append(pt2)
                
                #start turn
                pt1 = self.a.scatter(p[2], 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
                pickable_artists1.append(pt1)
                
                
            self.btn_detect_step["state"] = "normal"
            phases=np.vstack(self.phase)
            self.canvas.draw()
            self.turn_stop_ind=phases[:,1]
            self.turn_start_ind=phases[:,2]

            print("gyro_turn_period_hilbert")

            # low_idx, high_idx =self.hl_envelopes_idx(self.CP_data.gyro_magnitude, 
            #                                           dmin=1, dmax=1, split=False)
            
            

            # plt.figure()
            # plt.plot(high_idx,self.CP_data.gyro_magnitude[high_idx])
            # plt.plot(self.CP_data.gyro_magnitude)
            # plt.plot(amplitude_envelope_acc)

            #########################################################
            turns=self.window_rms(self.CP_data_phone.gyro_magnitude, window_size=300)
            peak_index,_= find_peaks(turns,distance=500)

            gyro_max=np.mean(turns[peak_index])
            
            # gyro_max=np.mean(turns[str_walk:stp_walk])
            print(gyro_max)
            gyro_binary=np.where(turns>gyro_max)[0]
            ranges=self.ranges(gyro_binary)
            turn_array=np.zeros(len(self.CP_data_phone.gyro_magnitude))

            for (start,end) in ranges:
                if start+50<end:
                    turn_array[start:end]=1
            
            #########################################################
            # walking_zone=[(self.walking_period[0],self.walking_period[1])]
            # accc,gyroo=self.CP_data.crop_medipole(fs=1,phases=walking_zone,turn_time=0)
            
            # self.CP_data.filter_data(acc=accc[0],gyro=gyroo[0],N=10,fc=12,fs=100)
            # self.CP_data.calculate_norm_accandgyro(gyro=self.CP_data.gyro_filtered,acc=self.CP_data.acc_filtered)

            # gyro_mag=self.CP_data.gyro_magnitude
            # acc_mag=self.CP_data.acc_magnitude
            # gyro_max=np.amax(gyro_mag)
            # acc_mean=np.mean(acc_mag)
            
            # self.CP_data.filter_data(acc=self.CP_data.acc_interp,gyro=self.CP_data.gyro_interp,N=10,fc=12,fs=100)
            # self.CP_data.calculate_norm_accandgyro(gyro=self.CP_data.gyro_filtered,acc=self.CP_data.acc_filtered)
            
            # gyro_mag=self.CP_data.gyro_magnitude
            # acc_mag=self.CP_data.acc_magnitude
            
            # gyro_binary=np.where(self.CP_data.gyro_magnitude>gyro_max)[0]
            # # acc_binary=[self.CP_data.acc_magnitude<acc_mean]
            # # gyro_binary=np.transpose(np.array(gyro_binary).astype("int"))
            
            # ranges=self.ranges(gyro_binary)
            
            # turn_array=np.zeros(len(gyro_mag))

            # for (start,end) in ranges:
            #     if start+50<end:
            #         turn_array[start:end]=1
            
            plt.figure()
            plt.plot(self.CP_data_phone.gyro_magnitude)
            plt.plot(turn_array)
            plt.plot(turns)
            
            self.turns=turn_array
            
            turn_array=turn_array+np.mean(self.CP_data_phone.acc_magnitude)
            

            self.a.plot(turn_array, color='red',alpha=0.5)
            self.b.plot(turn_array/5, color='red',alpha=0.5)
            if self.rightfoot or self.leftfoot:
                self.c.plot(turn_array/5, color='red',alpha=0.5)
            if self.hand:
                self.d.plot(turn_array, color='red',alpha=0.5)

 
            self.canvas.draw()
                    
                    
                    
            
    def window_rms(self,a, window_size):
        a2 = np.power(a,2)
        window = np.ones(window_size)/float(window_size)
        return np.sqrt(np.convolve(a2, window, 'same'))
            
            


    def hl_envelopes_idx(self,s, dmin=1, dmax=1, split=False):
        """
        Input :
        s: 1d-array, data signal from which to extract high and low envelopes
        dmin, dmax: int, optional, size of chunks, use this if the size of the input signal is too big
        split: bool, optional, if True, split the signal in half along its mean, might help to generate the envelope in some cases
        Output :
        lmin,lmax : high/low envelope idx of input signal s
        """
    
        # locals min      
        lmin = (np.diff(np.sign(np.diff(s))) > 0).nonzero()[0] + 1 
        # locals max
        lmax = (np.diff(np.sign(np.diff(s))) < 0).nonzero()[0] + 1 
        
    
        if split:
            # s_mid is zero if s centered around x-axis or more generally mean of signal
            s_mid = np.mean(s) 
            # pre-sorting of locals min based on relative position with respect to s_mid 
            lmin = lmin[s[lmin]<s_mid]
            # pre-sorting of local max based on relative position with respect to s_mid 
            lmax = lmax[s[lmax]>s_mid]
    
    
        # global max of dmax-chunks of locals max 
        lmin = lmin[[i+np.argmin(s[lmin[i:i+dmin]]) for i in range(0,len(lmin),dmin)]]
        # global min of dmin-chunks of locals min 
        lmax = lmax[[i+np.argmax(s[lmax[i:i+dmax]]) for i in range(0,len(lmax),dmax)]]
        
        return lmin,lmax
            
            
    def ranges(self,nums,d=200):
        nums = sorted(set(nums))
        gaps = [[s, e] for s, e in zip(nums, nums[1:]) if s+d < e]
        edges = iter(nums[:1] + sum(gaps, []) + nums[-1:])
        return list(zip(edges, edges))


    def extract_signal_walking(self):
        phases_adj=[]
        turn_strides=[]
        phases=[]
        for line in self.treeview.get_children():
            phases.append(self.treeview.item(line)['values'])
        if (len(self.treeview_res.get_children())!=0):
            self.treeview_res.delete(*self.treeview_res.get_children())
        if (len(self.treeview_stat.get_children())!=0):
            self.treeview_stat.delete(*self.treeview_stat.get_children())
        for (x,start,stop,z) in phases:
            start=int(start)
            stop=int(stop)
            z=int(z)
            phases_adj.append((start,stop))
            turn_strides.append(z)
        accc,gyroo=self.CP_data_phone.crop_medipole(fs=1,phases=phases_adj,turn_time=0)
        
        return(accc,gyroo)
        
    def Detect_steps(self,event=None):
        phases_adj=[]
        turn_strides=[]
        phases=[]
        tot_peaks=[]
        result_adj=[]
        n_phases=0
        for line in self.treeview.get_children():
            phases.append(self.treeview.item(line)['values'])
        
        
        if (len(self.treeview_res.get_children())!=0):
            self.treeview_res.delete(*self.treeview_res.get_children())
        
        if (len(self.treeview_stat.get_children())!=0):
            self.treeview_stat.delete(*self.treeview_stat.get_children())
            
        for (x,start,stop,z) in phases:
            start=int(start)
            stop=int(stop)
            z=int(z)
            phases_adj.append((start,stop))
            turn_strides.append(z)
            

        accc,gyroo=self.CP_data_phone.crop_medipole(fs=1,phases=phases_adj,turn_time=0)
        

        # df = pd.concat(accc,ignore_index=True)
        # plt.figure()
        # plt.plot(df)
        cum_ind=0     
        for i in range(0,len(accc)):
            
            if self.var_stepdetectionmethod.get()=="Use_smartstep":
                print("using smartstep")
            
                self.CP_data_phone.calculate_norm_accandgyro(gyro=gyroo[i],acc=accc[i])
                
                signals_unfiltered=[self.CP_data_phone.acc_magnitude,self.CP_data_phone.gyro_magnitude]
                
                self.CP_data_phone.filter_data(acc=accc[i],gyro=gyroo[i],N=10,fc=2,fs=100)
                
                self.CP_data_phone.calculate_norm_accandgyro(gyro=self.CP_data_phone.gyro_filtered,acc=self.CP_data_phone.acc_filtered)
                
                signals=[self.CP_data_phone.acc_magnitude,self.CP_data_phone.gyro_magnitude]
                
                gyro_features,acc_features=self.CP_data_phone.calculate_features(signals,signals_unfiltered)
        
                
                steps_in_window,step_type=self.CP_data_phone.predict_steps(gyro_features,acc_features)
                
                steps_smartstep=self.CP_data_phone.inverse_window_step(steps_in_window)
                
                steps_smartstep=np.array(steps_smartstep)
                
                cum_ind=phases_adj[i][0]
                peaks=steps_smartstep+cum_ind

            if self.var_stepdetectionmethod.get()=="Use_peak_detection": 
                
                print("using step_detection")
                
                self.CP_data_phone.filter_data(acc=accc[i],gyro=gyroo[i],N=10,fc=3,fs=100)
                self.CP_data_phone.calculate_norm_accandgyro(gyro=self.CP_data_phone.gyro_filtered,acc=self.CP_data_phone.acc_filtered)
                
                self.CP_data_phone.peakdet_m2(acc=True,plot_peak=False,detect_turn=False)
                
                peaks=self.CP_data_phone.peakandvalley["peak_index"]
                
                peaks=np.array(peaks)
                
                cum_ind=phases_adj[i][0]
                peaks=peaks+cum_ind

            if (len(peaks)>5):
                print(len(peaks))
                tot_peaks.append(peaks)
                self.CP_data_phone.computeVarStride(fs=100,remove_outliers=True,N=3,use_smartstep=True,manual_peaks=peaks,use_peaks=True,
                                              pocket=False,remove_step=turn_strides[i])
            
                try:
                    if len(self.CP_data_phone.cycle_temp['stridetime'])>5:
                        n_phases=n_phases+1
                        result_adj_phase=np.c_[ np.zeros(len(self.CP_data_phone.cycle_temp["detailed_stridetime"]))+i, 
                                               self.CP_data_phone.cycle_temp["detailed_stridetime"] ] 
                        
                        result_adj.append(result_adj_phase)

                        for d in result_adj_phase:
                            self.treeview_res.insert('', tk.END, values=(int(d[0]),int(d[1]),int(d[2]),d[3]))
                         
                        d=((i,np.around(np.mean(self.CP_data_phone.cycle_temp['stridetime']),decimals=3),
                            self.CP_data_phone.cycle_temp['stridetime_std'],
                            self.CP_data_phone.cycle_temp['stridetime_Cov'],
                            len(self.CP_data_phone.cycle_temp['stridetime']),len(accc[i])/100)) 
                        
                        self.treeview_stat.insert('',tk.END,values=d)
                except Exception as e:
                    print (repr(e))
                    print("no calculation of stride time for this segment")
            else:
                print("no steps occured in this walking period")
         
            
        self.tot_peaks=np.hstack(tot_peaks)
        result_adj=np.vstack(result_adj)

        exported_list=[]
        for line in self.treeview_stat.get_children():
            exported_list.append(self.treeview_stat.item(line)['values'])
            
        self.exported_results_summary_phone=pd.DataFrame(exported_list, columns=["Phase", "Mean stride duration", "standard deviation of stride duration","Coefficient of variance of stride duration", "Number of strides","Duration"])
        
            
        self.exported_results_phone=pd.DataFrame(result_adj, columns=["Phase","Start Heel strike foot",
                                                                      "Stop Heel strike foot",
                                                                      "stride duration"])
        
        self.lbl_mean_stridetime.configure(text ='%.2f seconds'%(np.round(np.mean(self.exported_results_phone['stride duration'].values),decimals=5)))
        self.lbl_SDstride.configure(text ='%.2f milliseconds'%(np.round(np.std(self.exported_results_phone['stride duration'].values),decimals=5)*1000))
        self.lbl_CVstride.configure(text ='%.2f %%'%(np.round(np.std(self.exported_results_phone['stride duration'].values)/np.mean(self.exported_results_phone['stride duration'].values),decimals=5)*100))
        self.lbl_Nstride.configure(text ='%d'%(len(self.exported_results_phone['stride duration'].values)))
        self.lbl_Nphase.configure(text ='%d'%(n_phases))
        
        if self.right_excel or self.left_excel:
            self.Calculate_metrics(position="waist")
            phone_steps=self.True_positives_phases_waist.iloc[:,2]
            gaitup_steps=self.True_positives_phases_waist.iloc[:,1]
            
            self.CP_data_phone.computeVarStride(fs=100,remove_outliers=False,N=3,use_smartstep=True,manual_peaks=phone_steps,use_peaks=True,
                                              pocket=False,remove_step=0)
            
            self.phone_strides=pd.DataFrame(self.CP_data_phone.cycle_temp["detailed_stridetime"],columns=["Heel_strike","Foot_off","Stride_time"])
            

            self.CP_data_phone.computeVarStride(fs=100,remove_outliers=False,N=3,use_smartstep=True,manual_peaks=gaitup_steps,use_peaks=True,
                                              pocket=False,remove_step=0)
            
            self.gaitup_strides=pd.DataFrame(self.CP_data_phone.cycle_temp["detailed_stridetime"],columns=["Heel_strike","Foot_off","Stride_time"])

            phone_index_todrop=self.phone_strides[(self.phone_strides.Stride_time > 3) | (self.phone_strides.Stride_time < 0) ].index
            
            self.phone_strides=self.phone_strides.drop(phone_index_todrop)
        
            self.gaitup_strides=self.gaitup_strides.drop(phone_index_todrop)
    
            gaitup_index_todrop=self.gaitup_strides[(self.gaitup_strides.Stride_time > 3) | (self.gaitup_strides.Stride_time < 0) ].index
    
            self.gaitup_strides=self.gaitup_strides.drop(gaitup_index_todrop)
            
            self.phone_strides=self.phone_strides.drop(gaitup_index_todrop)

            




        self.plot_stridetime()
        
    
    
    def Calculate_metrics(self,event=None,position="waist"):
        
        if position=="waist":
            i=0
            confusion_metrixs_phases=[]
            True_positives=[]
            Missdetected=[]
            for phase in self.exported_results_summary_phone["Phase"].values:
                
                index_phase=np.where(self.exported_results_phone["Phase"]==phase)[0].astype("int")
                
                
                steps_phone=np.concatenate((self.exported_results_phone["Start Heel strike foot"].values[index_phase],
                                           self.exported_results_phone["Stop Heel strike foot"].values[index_phase]))
                steps_phone.sort()
                steps_phone=np.unique(steps_phone)
                
                print(len(steps_phone))
    
                index_phase=np.where(self.exported_results_feet["Phase"]==phase)[0].astype("int")
                
                steps_gaitup=np.concatenate((self.exported_results_feet["Start Heel strike foot"].values[index_phase],
                                           self.exported_results_feet["Stop Heel strike foot"].values[index_phase]))
                
                steps_gaitup.sort()
                steps_gaitup=np.unique(steps_gaitup)
        
                
                confusion_matrix,TP,MD=self.CP_data_phone.calculate_metrics(all_steps=steps_gaitup,steps=steps_phone,len_gyro_feautres=int(float(self.exported_results_summary_phone["Duration"][i])*100))
                i=i+1
                TP=np.vstack(TP)
                TP=np.c_[np.zeros((len(TP),1)),TP]
                True_positives.append(TP)
                Missdetected.append(MD)
                confusion_metrixs_phases.append(confusion_matrix)
                
            confusion_metrixs_phases=np.vstack(confusion_metrixs_phases)
            True_positives_phases=np.vstack(True_positives)
            Missdetected_phases=np.transpose(np.hstack(Missdetected))
        
            self.confusion_matrixs=pd.DataFrame(confusion_metrixs_phases,columns=["Truenegative","Falsepostive",
                                                                              "Falsenegative","Truepositive"])
            
            start_true=[]
            stop_true=[]
            for i in range(0,len(True_positives_phases)):
                start_true.append(np.where(self.exported_results_phone["Start Heel strike foot"].values.astype('int')==int(True_positives_phases[i,1]))[0])
                stop_true.append(np.where(self.exported_results_phone["Stop Heel strike foot"].values.astype('int')==int(True_positives_phases[i,1]))[0])
                print(len(start_true))
                print(len(stop_true))
                if len(start_true)!=len(stop_true):
                    print("attention the highlighting not working properly")
                
            start_miss=[]
            stop_miss=[]
            for i in range(0,len(Missdetected_phases)):
                start_miss.append(np.where(self.exported_results_feet["Start Heel strike foot"].values==Missdetected_phases[i])[0])
                stop_miss.append(np.where(self.exported_results_feet["Stop Heel strike foot"].values==Missdetected_phases[i])[0])
                if len(start_miss)!=len(stop_miss):
                    print("attention the highlighting not working properly")   
            
            start_true=np.hstack(start_true)
            start_miss=np.hstack(start_miss)
            
            stop_true=np.hstack(stop_true)
            stop_miss=np.hstack(stop_miss)
            
            self.start=[start_true,start_miss]
            self.stop=[stop_true,stop_miss]
            
            self.True_positives_phases_waist=pd.DataFrame(True_positives_phases)

        if position=="hand":
            
            if self.hand:
                print("calculating metrics for the hand")
                i=0
                confusion_metrixs_phases=[]
                True_positives=[]
                Missdetected=[]
                
                for phase in self.exported_results_summary_hand["Phase"].values:
                    
                    index_phase=np.where(self.exported_results_hand["Phase"]==phase)[0].astype("int")
                    
                    steps_hand=np.concatenate((self.exported_results_hand["Start Heel strike foot"].values[index_phase],
                                           self.exported_results_hand["Stop Heel strike foot"].values[index_phase]))
                    steps_hand.sort()
                    steps_hand=np.unique(steps_hand)
                    
                    print(len(steps_hand))
                    
                    index_phase=np.where(self.exported_results_feet["Phase"]==phase)[0].astype("int")
                    
                    steps_gaitup=np.concatenate((self.exported_results_feet["Start Heel strike foot"].values[index_phase],
                                               self.exported_results_feet["Stop Heel strike foot"].values[index_phase]))
                    
                    steps_gaitup.sort()
                    steps_gaitup=np.unique(steps_gaitup)
                    
                    confusion_matrix,TP,MD=self.CP_data_h.calculate_metrics(all_steps=steps_gaitup,steps=steps_hand,len_gyro_feautres=int(float(self.exported_results_summary_hand["Duration"][i])*100))
                    i=i+1
                    
                    TP=np.vstack(TP)
                    TP=np.c_[np.zeros((len(TP),1)),TP]
                    True_positives.append(TP)
                    Missdetected.append(MD)
                    confusion_metrixs_phases.append(confusion_matrix)
                    
                    
                confusion_metrixs_phases=np.vstack(confusion_metrixs_phases)
                True_positives_phases=np.vstack(True_positives)
                Missdetected_phases=np.transpose(np.hstack(Missdetected))
            
                self.confusion_matrixs_hand=pd.DataFrame(confusion_metrixs_phases,columns=["Truenegative","Falsepostive",
                                                                                  "Falsenegative","Truepositive"])
                
                start_true=[]
                stop_true=[]
                for i in range(0,len(True_positives_phases)):
                    start_true.append(np.where(self.exported_results_hand["Start Heel strike foot"].values.astype('int')==int(True_positives_phases[i,1]))[0])
                    stop_true.append(np.where(self.exported_results_hand["Stop Heel strike foot"].values.astype('int')==int(True_positives_phases[i,1]))[0])
                    print(len(start_true))
                    print(len(stop_true))
                    if len(start_true)!=len(stop_true):
                        print("attention the highlighting not working properly")
                    
                start_miss=[]
                stop_miss=[]
                for i in range(0,len(Missdetected_phases)):
                    start_miss.append(np.where(self.exported_results_feet["Start Heel strike foot"].values==Missdetected_phases[i])[0])
                    stop_miss.append(np.where(self.exported_results_feet["Stop Heel strike foot"].values==Missdetected_phases[i])[0])
                    if len(start_miss)!=len(stop_miss):
                        print("attention the highlighting not working properly")   
                
                start_true=np.hstack(start_true)
                start_miss=np.hstack(start_miss)
                
                stop_true=np.hstack(stop_true)
                stop_miss=np.hstack(stop_miss)
                
                self.start_hand=[start_true,start_miss]
                self.stop_hand=[stop_true,stop_miss]
                
                
                
                self.True_positives_phases_hand=pd.DataFrame(True_positives_phases)
            
        

    def plot_and_calculate_start(self,event=None):
        print("doing")
        
        self.read_gaitup_phone_csv()
        
        # if self.leftfoot and self.rightfoot:    
        #     #start of walking in foot signal is the minimum from right and left
        #     if self.right_excel:
        #         s_rf=int((self.start_excel_right)*10000//128)
        #         s_f=s_rf
        #     if self.left_excel:
        #         s_lf=int((self.start_excel_left)*10000//128)
        #         s_f=s_lf
        #     if self.right_excel and self.left_excel:
        #         s_f=np.minimum(s_rf,s_lf)

        #     print(s_f)
        #     self.ent_str_gaitup_feet.delete(0,tk.END)
        #     self.ent_str_gaitup_feet.insert(0,s_f)
        
        self.plot_gaitup_phone()
        
        
        
    def read_gaitup_phone_csv(self):
        """
        read the csv files of gaitup: left foot, right foot, hand into a dataframe
        interpolates and transform into CP object, to filter data from hand
        interpolate to 100 Hz 

        Returns
        -------
        None.

        """
        gaitup_dir=os.path.join(self.path,"Gaitup" )
        self.hand=True
        self.leftfoot=True
        self.rightfoot=True
        #hand norm
        filename=os.path.join(gaitup_dir,"hand.csv" )
        col_list = ["Time", "Gyro X","Gyro Y","Gyro Z","Accel X", "Accel Y","Accel Z"]
        try:
            df_h=pd.read_csv(filename, delimiter=",",skiprows=[0],usecols=col_list)

        except:
            self.hand=False
            print("no hand file")
            

            
        if self.hand:
            df_h = df_h.iloc[1:]
            df_h=df_h.astype('float32')
            
            self.df_h=df_h
            time=self.df_h['Time'].values
            
            #--interpolate manually--#
            self.df_h=self.df_h.set_index('Time')
            fs=100
            t_t=np.linspace(0, time[-1], num=np.int(time[-1]*fs), 
                            endpoint=True,dtype=np.float32)
            self.df_h=self.df_h.reindex(self.df_h.index.union(t_t))
            self.df_h=self.df_h.interpolate(method='linear', limit_direction='both',
                                            axis=0)
            self.df_h=self.df_h[self.df_h.index.isin(pd.Index(t_t))]
            self.df_h.index=np.around(self.df_h.index.astype('float32'),decimals=4)
            time =self.df_h.index
            acc_h=self.df_h.iloc[:,3:6]
            acc_h=acc_h.mul(9.8)
            gyro_h=self.df_h.iloc[:,0:3]
            
            
            #--transform into CP-object--#
            self.CP_data_h=CP(acc=acc_h,gyro=gyro_h,app="manual_entry")
            self.CP_data_h.deg2rad()
            self.CP_data_h.filter_data(acc=self.CP_data_h.acc_interp,gyro=self.CP_data_h.gyro_interp,N=10,fc=2,fs=100)
            self.CP_data_h.calculate_norm_accandgyro(gyro=self.CP_data_h.gyro_filtered,acc=self.CP_data_h.acc_filtered)
            self.acc_h_magnitude=self.CP_data_h.acc_magnitude
            self.gyro_h_magnitude=self.CP_data_h.gyro_magnitude
            
            #--Detect start of walk--#
            s_h=self.CP_data_h.detectstartofwalk(sig1=self.CP_data_h.acc_magnitude,thresh=3)

            self.ent_str_gaitup_hand.delete(0,tk.END)
            self.ent_str_gaitup_hand.insert(0,s_h)
        
        #--read left foot csv--#
        filename=os.path.join(gaitup_dir,"left_foot.csv" ) 
        
        try:
            col_list = ["Time", "Gyro X","Gyro Y","Gyro Z","Accel X", "Accel Y","Accel Z"]
            df_lf=pd.read_csv(filename, delimiter=",",skiprows=[0],usecols=col_list)

        except:
            self.leftfoot=False
            print("no leftfoot file")
        
        if self.leftfoot:
            df_lf = df_lf.iloc[1:]
            df_lf=df_lf.astype('float32')
            
            self.df_lf=df_lf
            time=self.df_lf['Time'].values
            #interpolate and filter
            self.df_lf=self.df_lf.set_index('Time')
            fs=100
            t_t=np.linspace(0, time[-1], num=np.int(time[-1]*fs), 
                            endpoint=True,dtype=np.float32)
    
            self.df_lf=self.df_lf.reindex(self.df_lf.index.union(t_t))
            self.df_lf=self.df_lf.interpolate(method='linear', limit_direction='both',
                                            axis=0)
            self.df_lf=self.df_lf[self.df_lf.index.isin(pd.Index(t_t))]
    
            self.df_lf.index=np.around(self.df_lf.index.astype('float32'),decimals=4)
    
            time =self.df_lf.index
            
            acc_lf=self.df_lf.iloc[:,3:6]
            gyro_lf=self.df_lf.iloc[:,0:3]
            
            #--create cp object from csv of foot--#
            self.CP_data_lf=CP(acc=acc_lf,gyro=gyro_lf,app="manual_entry")
            self.CP_data_lf.calculate_norm_accandgyro(gyro=self.CP_data_lf.gyro_interp,acc=self.CP_data_lf.acc_interp)
            self.acc_lf_magnitude=self.CP_data_lf.acc_magnitude
            self.gyro_lf_magnitude=self.CP_data_lf.gyro_magnitude
            s_lf=self.CP_data_lf.detectstartofwalk(sig1=self.CP_data_lf.acc_magnitude,thresh=5)
            s_f=s_lf
        #--read right foot csv--#
        filename=os.path.join(gaitup_dir,"right_foot.csv" )
        
        try:
            col_list = ["Time", "Gyro X","Gyro Y","Gyro Z","Accel X", "Accel Y","Accel Z"]
            df_rf=pd.read_csv(filename, delimiter=",",skiprows=[0],usecols=col_list)

        except:
            self.rightfoot=False
            print("no rightfoot file")
        
        if self.rightfoot:
            #right foot
            df_rf = df_rf.iloc[1:]
            df_rf=df_rf.astype('float32')
            
            self.df_rf=df_rf
            time=self.df_rf['Time'].values
            #interpolate and filter
            self.df_rf=self.df_rf.set_index('Time')
            fs=100
            t_t=np.linspace(0, time[-1], num=np.int(time[-1]*fs), 
                            endpoint=True,dtype=np.float32)
    
            self.df_rf=self.df_rf.reindex(self.df_rf.index.union(t_t))
            self.df_rf=self.df_rf.interpolate(method='linear', limit_direction='both',
                                            axis=0)
            self.df_rf=self.df_rf[self.df_rf.index.isin(pd.Index(t_t))]
    
            self.df_rf.index=np.around(self.df_rf.index.astype('float32'),decimals=4)
    
            time =self.df_rf.index
            
            acc_rf=self.df_rf.iloc[:,3:6]
            gyro_rf=self.df_rf.iloc[:,0:3]
            # read right into cp object
            self.CP_data_rf=CP(acc=acc_rf,gyro=gyro_rf,app="manual_entry")
            self.CP_data_rf.calculate_norm_accandgyro(gyro=self.CP_data_rf.gyro_interp,acc=self.CP_data_rf.acc_interp)
            self.acc_rf_magnitude=self.CP_data_rf.acc_magnitude
            self.gyro_rf_magnitude=self.CP_data_rf.gyro_magnitude
            s_rf=self.CP_data_rf.detectstartofwalk(sig1=self.CP_data_rf.acc_magnitude,thresh=5)
            s_f=s_rf
            
        if self.leftfoot and self.rightfoot:    
            #start of walking in foot signal is the minimum from right and left
            s_f=np.minimum(s_rf,s_lf)
            self.ent_str_gaitup_feet.delete(0,tk.END)
            self.ent_str_gaitup_feet.insert(0,s_f)
            
            self.ent_str_gaitup_feet.delete(0,tk.END)
            self.ent_str_gaitup_feet.insert(0,s_f)
            
        
        
        
        #read phone data and detect start#
        self.CP_data_phone=CP(self.phone_direct,app="geoloc")
        self.CP_data_phone.interpolategyrnacc(fs=100)
        self.CP_data_phone.filter_data(acc=self.CP_data_phone.acc_interp,gyro=self.CP_data_phone.gyro_interp,N=10,fc=3,fs=100)
        self.CP_data_phone.calculate_norm_accandgyro(gyro=self.CP_data_phone.gyro_filtered,acc=self.CP_data_phone.acc_filtered)
        self.acc_w_magnitude=self.CP_data_phone.acc_magnitude
        
        self.s_w=self.CP_data_phone.detectstartofwalk(sig1=self.CP_data_phone.acc_magnitude,thresh=2)
        self.ent_str_phone_waist.delete(0,tk.END)
        self.ent_str_phone_waist.insert(0,self.s_w)
        self.e_w=int(len(self.CP_data_phone.acc_magnitude)-self.CP_data_phone.detectstartofwalk(sig1=self.CP_data_phone.acc_magnitude[::-1]))
        print("last walk inverse")
        print(self.CP_data_phone.detectstartofwalk(sig1=self.CP_data_phone.acc_magnitude[::-1]))
        
        print("length of signal")
        print(len(self.CP_data_phone.acc_magnitude))
        print("end of walk")
        print(self.e_w)
        self.btn_detect_start_stop["state"] = "normal"
        
        self.ent_str_walk.delete(0,tk.END)
        self.ent_str_walk.insert(0,self.s_w)
        self.ent_stp_walk.delete(0,tk.END)
        self.ent_stp_walk.insert(0,self.e_w)
        self.btn_detect_turn["state"] = "normal"
        

    def plot_gaitup_phone(self):
        self.figure1 = Figure(figsize=(4, 3), dpi=100)
        self.canvas1 = FigureCanvasTkAgg(self.figure1, master=self.fcontainer_gaituphand)
        self.canvas1.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas1, self.fcontainer_gaituphand)
        self.toolbar.update()
        self.canvas1._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        color = 'tab:blue'
        
        if self.hand:
            b = self.figure1.add_subplot(111)
            b.plot(self.acc_h_magnitude,zorder=1)
            b.set_ylabel('Acc m/s\u00b2', color=color)
            b.title.set_text("Hand")
            b.set_xticks([])
            s_h=int((self.ent_str_gaitup_hand.get()))
            b.scatter(s_h,self.acc_h_magnitude[s_h],color='red',marker="|",s=10000000000,zorder=2)
            b.set(xlim=(s_h-1000, s_h+1000))
            self.canvas1.draw()
        
        self.figure2 = Figure(figsize=(4, 3), dpi=100)
        self.canvas2 = FigureCanvasTkAgg(self.figure2, master=self.fcontainer_gaitupfeet)
        self.canvas2.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas2, self.fcontainer_gaitupfeet)
        self.toolbar.update()
        self.canvas2._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        color = 'tab:blue'
        
        b = self.figure2.add_subplot(111)
        if self.rightfoot:
            b.plot(self.acc_rf_magnitude,zorder=1)
        if self.leftfoot:
            b.plot(self.acc_lf_magnitude,zorder=1)

        s_f=int((self.ent_str_gaitup_feet.get()))
        if self.leftfoot:
            b.scatter(s_f,self.acc_lf_magnitude[s_f], color='red',marker="|",s=10000000000,zorder=2)
            b.set_ylabel('Acc m/s\u00b2', color=color)
            b.title.set_text("feet")
            b.set_xticks([])
            b.set(xlim=(s_f-1000, s_f+1000))
        elif self.rightfoot:
            b.scatter(s_f,self.acc_rf_magnitude[s_f], color='red',marker="|",s=10000000000,zorder=2)
            b.set_ylabel('Acc m/s\u00b2', color=color)
            b.title.set_text("feet")
            b.set_xticks([])
            b.set(xlim=(s_f-1000, s_f+1000))
            
        
        self.canvas2.draw()
        
        self.figure3 = Figure(figsize=(4, 3), dpi=100)
        self.canvas3 = FigureCanvasTkAgg(self.figure3, master=self.fcontainer_phonewaist)
        self.canvas3.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas3, self.fcontainer_phonewaist)
        self.toolbar.update()
        self.canvas3._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        color = 'tab:blue'
        b = self.figure3.add_subplot(111)
        b.plot(self.acc_w_magnitude,zorder=1)
        b.set_ylabel('Acc m/s\u00b2', color=color)
        b.title.set_text("waist")
        b.set_xticks([])
        s_w=int((self.ent_str_phone_waist.get()))
        b.scatter(s_w,self.acc_w_magnitude[s_w],color='red',marker="|",s=10000000000,zorder=2)
        b.set(xlim=(s_w-1000, s_w+1000))
        
        self.canvas3.draw()
  
    def read_gaitup_excel(self,event=None):
        self.right_excel=False
        self.left_excel=False
        gaitup_dir=os.path.join(self.path,"Gaitup_turn" )
        for root, dirs, files in os.walk(gaitup_dir):
            for file in files:
                if file.endswith(".xlsx"):
                    if 'right' in file:
                        right_step_file= file
                        self.right_excel=True
                        print(file)
                    if 'left' in file:
                        left_step_file= file
                        self.left_excel=True
                        print(file)
        if self.right_excel:
            #right 
            right_file=os.path.join(gaitup_dir,right_step_file)
            
            
            wb_right= load_workbook(right_file)
            ws_right=wb_right["Sheet1"]
            
            start_analysis_gaitup_right = ws_right['B14'].value
            stop_analysis_gaitup_right=ws_right['D14'].value
            
            ColD_right=ws_right['D']
            ColD_right=ColD_right[25::]
            
            turns_HS_right=[]
            straightwalk_HS_right=[]
            i=0
            for cl in ColD_right:
                if cl.value!= None:
                    i=i+1
                    if cl.font.color != None and type(cl.font.color.rgb) == str:
                        x=cl.value+start_analysis_gaitup_right
                        turns_HS_right.append([i,x])
                    else:
                        x=cl.value+start_analysis_gaitup_right
                        straightwalk_HS_right.append([i,x])
                    
            turns_HS_right=np.vstack(turns_HS_right)
            straightwalk_HS_right=np.vstack(straightwalk_HS_right)
            
            #get index of straight walking steps in the main csv signal
            time=self.CP_data_rf.acc_interp.index.values
            straightwalk_HS_right_index=[]
            for hs in straightwalk_HS_right[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    straightwalk_HS_right_index.append(t1)
                else:
                    straightwalk_HS_right_index.append(t2)
            
            #get index of turn walking steps in the main csv signal
            Turn_HS_right_index=[]
            for hs in turns_HS_right[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    Turn_HS_right_index.append(t1)
                else:
                    Turn_HS_right_index.append(t2)
                    
            self.HS_r_index=np.array(straightwalk_HS_right_index)
            self.start_excel_right=start_analysis_gaitup_right
                    
        if self.left_excel:
                
            #left
            left_file=os.path.join(gaitup_dir,left_step_file)
            
            
            wb_left= load_workbook(left_file)
            ws_left=wb_left["Sheet1"]
            
            start_analysis_gaitup_left = ws_left['B14'].value
            stop_analysis_gaitup_left=ws_left['D14'].value
            
            ColD_left=ws_left['D']
            ColD_left=ColD_left[25::]
            
            turns_HS_left=[]
            straightwalk_HS_left=[]
            i=0
            for cl in ColD_left:
                if cl.value!= None:
                    i=i+1
                    if cl.font.color != None and type(cl.font.color.rgb) == str:
                        x=cl.value+start_analysis_gaitup_left
                        turns_HS_left.append([i,x])
                    else:
                        x=cl.value+start_analysis_gaitup_left
                        straightwalk_HS_left.append([i,x])
                    
            turns_HS_left=np.vstack(turns_HS_left)
            straightwalk_HS_left=np.vstack(straightwalk_HS_left)
            
            #get index of straight walking steps in the main csv signal
            time=self.CP_data_lf.acc_interp.index.values
            straightwalk_HS_left_index=[]
            for hs in straightwalk_HS_left[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    straightwalk_HS_left_index.append(t1)
                else:
                    straightwalk_HS_left_index.append(t2)
            
            #get index of turn walking steps in the main csv signal
            Turn_HS_left_index=[]
            for hs in turns_HS_left[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    Turn_HS_left_index.append(t1)
                else:
                    Turn_HS_left_index.append(t2)
                    
            self.HS_l_index=np.array(straightwalk_HS_left_index)
            self.start_excel_left=start_analysis_gaitup_left   
        # straight_walking_periods_left=self.ranges(straightwalk_HS_left[:,0])
        
        # turning_periods_left=self.ranges(turns_HS_left[:,0],d=2)
        
        if self.left_excel and self.right_excel:
            turn_right_left=np.concatenate((turns_HS_left[:,1],turns_HS_right[:,1]))
        
        elif self.left_excel:
            turn_right_left=turns_HS_left[:,1]
        
        elif self.right_excel:
            turn_right_left=turns_HS_right[:,1]
            
        if self.left_excel or self.right_excel:
            turn_right_left=turn_right_left*100
            turn_right_left=turn_right_left.astype("int")
            turn_right_left.sort()
            turn_right_left=np.unique(turn_right_left)
            
            self.turning_periods_left_right=self.ranges(turn_right_left,d=400)
            self.turning_periods_left_right=np.array(self.turning_periods_left_right)
            self.turning_periods_left_right=np.vstack(self.turning_periods_left_right)
            print(self.turning_periods_left_right)
        
        #removing dupliactes in case turn is last step

            i=0
            while i<len(self.turning_periods_left_right):
                if self.turning_periods_left_right[i,0]==self.turning_periods_left_right[i,1]:
                    self.turning_periods_left_right=np.delete(self.turning_periods_left_right,i,0)
                i=i+1
    
            print(self.turning_periods_left_right)
            

        

        
    def crop_signals(self,event=None,end="shortest"):
        
        s_h=int((self.ent_str_gaitup_hand.get()))
        s_w=int((self.ent_str_phone_waist.get()))
        s_f=int((self.ent_str_gaitup_feet.get()))
        stop_phone=int((self.ent_stp_walk.get()))-s_w
        
        try:
            f = open(self.path+"\\synch.txt", "r")
            n=f.readline().split(",")
            print(n)
            if n[0]!='':
                s_f=int(n[0])
                self.ent_str_gaitup_feet.delete(0,tk.END)
                self.ent_str_gaitup_feet.insert(0,s_f)
                
            if n[1]!='':
                s_h=int(n[1])    
                self.ent_str_gaitup_hand.delete(0,tk.END)
                self.ent_str_gaitup_hand.insert(0,s_h)
            if n[2]!='':
                s_w=int(n[2])
                self.ent_str_phone_waist.delete(0,tk.END)
                self.ent_str_phone_waist.insert(0,s_w)
                stop_phone=int((self.ent_stp_walk.get()))-s_w
            print("start and stop taken from text file")
        except:
            print("text file of synch doesnt exist")
        


        stop_ind_feet=stop_phone
        print(s_h)
        
        if self.rightfoot:
            self.CP_data_rf.manual_crop(ind_start=s_f)
            
        if self.leftfoot:
            self.CP_data_lf.manual_crop(ind_start=s_f)
            
        if self.left_excel or self.right_excel:
            print("aligning turns start")
            print(s_f)
            self.turning_periods_left_right=self.turning_periods_left_right-s_f
        
        if self.right_excel:
            self.HS_r_index=self.HS_r_index-s_f
            stop_ind_feet=self.HS_r_index[-1]
            self.e_g=stop_ind_feet
            
        if self.left_excel:
            self.HS_l_index=self.HS_l_index-s_f
            stop_ind_feet=self.HS_l_index[-1]
            self.e_g=stop_ind_feet
            
        if self.left_excel and self.right_excel:
            stop_ind_feet=np.maximum(self.HS_l_index[-1],self.HS_r_index[-1])+1
            self.e_g=stop_ind_feet

            
        self.CP_data_phone.manual_crop(ind_start=s_w)
        if self.hand:
            print("hand is available")
            self.CP_data_h.manual_crop(ind_start=s_h)
            
        self.acquisition_stops=False
        if end=="shortest":
            print("shortest")
            print(stop_ind_feet)
            if self.left_excel or self.right_excel:
                if stop_phone<stop_ind_feet-1000:
                    print("the phone stops acquisition before end of trial")
                    self.acquisition_stops=True
                stop_ind_feet=np.minimum(stop_phone,stop_ind_feet)
                print(stop_ind_feet)
                
                ind_turn=np.where((self.turning_periods_left_right[:,0]<=stop_ind_feet)&(self.turning_periods_left_right[:,1]<=stop_ind_feet))[0]
                self.turning_periods_left_right=self.turning_periods_left_right[ind_turn,:]
                
                if self.right_excel:
                    ind_foot=np.where(self.HS_r_index<=stop_ind_feet)[0]
                    self.HS_r_index=self.HS_r_index[ind_foot]
                if self.left_excel:
                    ind_foot=np.where(self.HS_l_index<=stop_ind_feet)[0]
                    self.HS_l_index=self.HS_l_index[ind_foot]

        #manual crop the interp#
        if self.rightfoot:
            self.CP_data_rf.manual_crop(ind_start=0,ind_stop=stop_ind_feet)
        if self.leftfoot:
            self.CP_data_lf.manual_crop(ind_start=0,ind_stop=stop_ind_feet)
        self.CP_data_phone.manual_crop(ind_start=0,ind_stop=stop_ind_feet)
        if self.hand:
            self.CP_data_h.manual_crop(ind_start=0,ind_stop=stop_ind_feet)

    def recalculate_filter_norm(self):
        
        if self.leftfoot:
            self.CP_data_lf.calculate_norm_accandgyro(gyro=self.CP_data_lf.gyro_interp,
                                                  acc=self.CP_data_lf.acc_interp)
            self.acc_lf_magnitude=self.CP_data_lf.acc_magnitude
        if self.rightfoot:
            self.CP_data_rf.calculate_norm_accandgyro(gyro=self.CP_data_rf.gyro_interp,
                                                  acc=self.CP_data_rf.acc_interp)
            self.acc_rf_magnitude=self.CP_data_rf.acc_magnitude
        
        if self.hand:
            self.CP_data_h.filter_data(acc=self.CP_data_h.acc_interp,
                                       gyro=self.CP_data_h.gyro_interp)
            self.CP_data_h.calculate_norm_accandgyro(gyro=self.CP_data_h.gyro_filtered,
                                                     acc=self.CP_data_h.acc_filtered)
            self.acc_h_magnitude=self.CP_data_h.acc_magnitude
            self.gyro_h_magnitude=self.CP_data_h.gyro_magnitude

        self.CP_data_phone.filter_data(acc=self.CP_data_phone.acc_interp,
                                       gyro=self.CP_data_phone.gyro_interp)
        self.CP_data_phone.calculate_norm_accandgyro(gyro=self.CP_data_phone.gyro_filtered,
                                                     acc=self.CP_data_phone.acc_filtered)
        self.acc_w_magnitude=self.CP_data_phone.acc_magnitude
        
    def align_signals(self,event=None):
        print("aligning")
        self.read_gaitup_excel(event=None)
        self.crop_signals()
        self.recalculate_filter_norm()
        self.plot_aligned_signals()
        self.Detect_turns()
        
    def plot_aligned_signals(self):

        self.figure4 = Figure(figsize=(4, 3), dpi=100)
        self.canvas4 = FigureCanvasTkAgg(self.figure4, master=self.fcontainer_allsignals)
        self.canvas4.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas4, self.fcontainer_allsignals)
        self.toolbar.update()
        self.canvas4._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        
        color = 'tab:blue'
        b = self.figure4.add_subplot(111)
        b.plot(self.acc_w_magnitude,zorder=1)
        if self.hand:
            b.plot(self.acc_h_magnitude,zorder=1)
            b.plot(self.gyro_h_magnitude,zorder=1)
        if self.leftfoot:    
            b.plot(self.acc_lf_magnitude,zorder=1)
        if self.rightfoot:
            b.plot(self.acc_rf_magnitude,zorder=1)
        
        b.set_ylabel('Acc m/s\u00b2', color=color)
        b.title.set_text("allsignals")
        b.set_xticks([])
        
        self.canvas4.draw()
        
        self.plot_graph(plot_name="x")
    
        self.calculate_gaitup_foot()
        
    def calculate_gaitup_foot(self):
        print("calculating gaitup feet")
        result_adj=[]
        
        if self.right_excel or self.left_excel:
        
            if (len(self.treeview_stat_GU.get_children())!=0):
                self.treeview_stat_GU.delete(*self.treeview_stat.get_children())
            
            stop_walk=self.turning_periods_left_right[:,0]
            start_walk=self.turning_periods_left_right[:,1]
            
            stop_walk=np.array([*stop_walk, len(self.acc_rf_magnitude)])
            start_walk=np.array([0, *start_walk])
            n_phases=0
            
            for i in range(0,len(stop_walk)):
                
                if self.left_excel:
                    steps_left_phase=self.HS_l_index[np.where((self.HS_l_index<stop_walk[i]) & (self.HS_l_index>start_walk[i]))[0]]
                    steps_all=steps_left_phase
                    
                    
                if self.right_excel:
                    steps_right_phase=self.HS_r_index[np.where((self.HS_r_index<stop_walk[i]) & (self.HS_r_index>start_walk[i]))[0]]
                    steps_all=steps_right_phase
                    
                if self.left_excel and self.right_excel:
                    steps_all=np.concatenate((steps_left_phase,steps_right_phase))
                    steps_all.sort()
                
                if (len(steps_all)>3):
                    print("not empty steps")
                    print("number of steps %d"%len(steps_all))
                    if self.left_excel and self.right_excel:
                        print("using 2 excel  files")
                        self.CP_data_lf.computeVarStride(fs=100,remove_outliers=True,N=3,use_smartstep=True,manual_peaks=steps_all,use_peaks=True,pocket=False,remove_step=0)
                    else:
                        print("using 1 excel  file")
                        # here pocket is true
                        self.CP_data_lf.computeVarStride(fs=100,remove_outliers=True,N=3,use_smartstep=True,manual_peaks=steps_all,use_peaks=True,pocket=True,remove_step=0)

                    try:
                        n_phases=n_phases+1
                        result_adj_phase=np.c_[ np.zeros(len(self.CP_data_lf.cycle_temp["detailed_stridetime"]))+i, 
                                               self.CP_data_lf.cycle_temp["detailed_stridetime"] ] 
                        
                        result_adj.append(result_adj_phase)
                        
                        # stride_times.append(self.CP_data_lf.cycle_temp['stridetime'])
                        try:
                            d=((i,np.around(np.mean(self.CP_data_lf.cycle_temp['stridetime']),decimals=3),
                                self.CP_data_lf.cycle_temp['stridetime_std'],
                                self.CP_data_lf.cycle_temp['stridetime_Cov'],
                                len(self.CP_data_lf.cycle_temp["stridetime"]),
                                (stop_walk[i]-start_walk[i])/100,(start_walk[i+1]-stop_walk[i])/100))
                        except:
                            d=((i,np.around(np.mean(self.CP_data_lf.cycle_temp['stridetime']),decimals=3),
                                self.CP_data_lf.cycle_temp['stridetime_std'],
                                self.CP_data_lf.cycle_temp['stridetime_Cov'],
                                len(self.CP_data_lf.cycle_temp["stridetime"]),
                                (stop_walk[i]-start_walk[i])/100,0))
                            
                        
                        self.treeview_stat_GU.insert('',tk.END,values=d)
                        
                    except Exception:
                        print("stridetimes are filtered in this region so it will be omitted")
                    
                    
                else:
                    print("no steps has occured in this period")
                              
            result_adj=np.vstack(result_adj)
            
            exported_list_feet=[]
            for line in self.treeview_stat_GU.get_children():
                exported_list_feet.append(self.treeview_stat_GU.item(line)['values'])
                
            self.exported_results_summary_feet=pd.DataFrame(exported_list_feet, 
                                                            columns=["Phase", "Mean stride duration", 
                                                                     "standard deviation of stride duration",
                                                                     "Coefficient of variance of stride duration",
                                                                     "Number of strides","Duration","Turn time"])

            self.exported_results_feet=pd.DataFrame(result_adj, columns=["Phase","Start Heel strike foot",
                                                                      "Stop Heel strike foot",
                                                                      "stride duration"])     
            
            self.lbl_mean_stridetime_GU.configure(text ='%.2f seconds'%(np.round(np.mean(result_adj[:,3]),decimals=5)))
            self.lbl_SDstride_GU.configure(text ='%.2f milliseconds'%(np.round(np.std(result_adj[:,3]),decimals=5)*1000))
            self.lbl_CVstride_GU.configure(text ='%.2f %%'%(np.round(np.std(result_adj[:,3])/np.mean(result_adj[:,3]),decimals=5)*100))
            self.lbl_Nstride_GU.configure(text ='%d'%(len(result_adj[:,3])))
            self.lbl_Nphase_GU.configure(text ='%d'%(n_phases))

    def plot_steps_onalignedsignals(self,event=None):

        self.figure6 = Figure(figsize=(5, 4), dpi=100)
        self.canvas6 = FigureCanvasTkAgg(self.figure6, master=self.fcontainer_signalwithstep)
        self.canvas6.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas6, self.fcontainer_signalwithstep)
        self.toolbar.update()
        self.canvas6._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        
        color = 'tab:blue'
       
        b = self.figure6.add_subplot(311)
        
        
        d=self.tot_peaks
        print(d)
        
        b.plot(self.acc_w_magnitude,zorder=1)
        
        self.HS_l_index=np.array(self.HS_l_index)
        self.HS_r_index=np.array(self.HS_r_index)
        print(self.HS_l_index)
        print(self.HS_r_index)
        st_l=self.HS_l_index[np.where(self.HS_l_index<len(self.acc_w_magnitude))[0]]
        st_r=self.HS_r_index[np.where(self.HS_r_index<len(self.acc_w_magnitude))[0]]
        
        b.scatter(st_l,self.acc_w_magnitude[st_l],color='red',marker="o",zorder=2)
        b.scatter(st_r,self.acc_w_magnitude[st_r],color='green',marker="o",zorder=2)

        #remove steps that are greater than the length of the signal
        d_adjusted=[step for step in d if step<len(self.acc_w_magnitude)-1]
        print("the number of steps removed because of gaitup crop reasons")
        print(len(d)-len(d_adjusted))
        
        b.scatter(d_adjusted,self.acc_w_magnitude[d_adjusted],color='black')
        b.set_ylabel('Acc m/s\u00b2', color=color)
        b.title.set_text("Gauche")
        b.set_xticks([])

        if self.hand:
            a = self.figure6.add_subplot(312)
            a.plot(self.acc_h_magnitude,zorder=1)
            a.scatter(self.HS_l_index,self.acc_h_magnitude[self.HS_l_index],color='red',marker="|",s=10000000000,zorder=2)
            a.scatter(self.HS_r_index,self.acc_h_magnitude[self.HS_r_index],color='green',marker="|",s=10000000000,zorder=2)
    
            a.set_ylabel('Acc m/s\u00b2', color=color)
            a.set_xlabel('Time', color=color)
            a.title.set_text("Hand")
            
            self.a=a
        
        c = self.figure6.add_subplot(313)
            
        c.plot(self.acc_lf_magnitude,zorder=1)
        c.plot(self.acc_rf_magnitude,zorder=1)
        
        c.scatter(self.HS_l_index,self.acc_lf_magnitude[self.HS_l_index],color='red',marker="|",s=10000000000,zorder=2)
        c.scatter(self.HS_r_index,self.acc_rf_magnitude[self.HS_r_index],color='green',marker="|",s=10000000000,zorder=2)

        
        c.set_ylabel('Acc m/s\u00b2', color=color)
        c.set_xlabel('Time', color=color)
        c.title.set_text("Right/Left foot")
        
        self.canvas1.draw()
        
        
    def detect_steps_hand(self,event=None):
        if self.hand:
            phases_adj=[]
            turn_strides=[]
            phases=[]
            tot_peaks=[]
            result_adj=[]
            
            #delete values already present in hand result treeview
            if (len(self.treeview_stat_HP.get_children())!=0):
                self.treeview_stat.delete(*self.treeview_stat_HP.get_children())
            
            # get turns from treeview
            for line in self.treeview.get_children():
                phases.append(self.treeview.item(line)['values'])
    
            for (x,start,stop,z) in phases:
                start=int(start)
                stop=int(stop)
                z=int(z)
                phases_adj.append((start,stop))
                turn_strides.append(z)
                
            #crop the signals
            accc,gyroo=self.CP_data_h.crop_medipole(fs=1,phases=phases_adj,turn_time=0)
            n_phases=0
            #step detection only works with smartstep
            for i in range(0,len(accc)):
                self.CP_data_h.calculate_norm_accandgyro(acc=accc[i],gyro=gyroo[i])
                
                signals_unfiltered=[self.CP_data_h.acc_magnitude,
                                self.CP_data_h.gyro_magnitude]
                
                self.CP_data_h.filter_data(acc=accc[i],gyro=gyroo[i])
                
                self.CP_data_h.calculate_norm_accandgyro(gyro=self.CP_data_h.gyro_filtered,
                                                         acc=self.CP_data_h.acc_filtered)
                
                signals=[self.CP_data_h.acc_magnitude,self.CP_data_h.gyro_magnitude]
                
                # plt.figure()
                # plt.plot(self.CP_data_h.acc_magnitude)
                # plt.plot(self.CP_data_h.gyro_magnitude)

                gyro_features,acc_features=self.CP_data_h.calculate_features(signals,
                                                                             signals_unfiltered)
                
                steps_in_window,step_type=self.CP_data_h.predict_steps(gyro_features,acc_features)
                
                steps_smartstep=self.CP_data_h.inverse_window_step(steps_in_window)
                
                steps_smartstep=np.array(steps_smartstep)

                
                if (len(steps_smartstep)>5):
                    cum_ind=phases_adj[i][0]
                    peaks=steps_smartstep+cum_ind
                    tot_peaks.append(peaks)

                    self.CP_data_h.computeVarStride(fs=100,remove_outliers=True,N=3,use_smartstep=True,manual_peaks=peaks,use_peaks=True,
                                              pocket=False,remove_step=turn_strides[i])

                    try:
                        if len(self.CP_data_h.cycle_temp['stridetime'])>5:
                            n_phases=n_phases+1
                            result_adj_phase=np.c_[ np.zeros(len(self.CP_data_h.cycle_temp["detailed_stridetime"]))+i, 
                                                   self.CP_data_h.cycle_temp["detailed_stridetime"] ] 
                            
                            result_adj.append(result_adj_phase)
                            
                            
                        d=((i,np.around(np.mean(self.CP_data_h.cycle_temp['stridetime']),decimals=3),
                            self.CP_data_h.cycle_temp['stridetime_std'],
                            self.CP_data_h.cycle_temp['stridetime_Cov'],
                            len(self.CP_data_h.cycle_temp["stridetime"]),len(accc[i])/100)) 
                        
                        self.treeview_stat_HP.insert('',tk.END,values=d)
                    
                    except Exception as e:
                        print (repr(e))
                        print("no calculation of stride time for this segment")
                    
                else:
                    print("very few steps were recorded by smartstep")
                   
                    
                   
                
            self.tot_peaks=np.hstack(tot_peaks)
            result_adj=np.vstack(result_adj)
            # add phase 
             
            
            exported_list_hand=[]
            for line in self.treeview_stat_HP.get_children():
                exported_list_hand.append(self.treeview_stat_HP.item(line)['values'])
                
            self.exported_results_summary_hand=pd.DataFrame(exported_list_hand, 
                                                            columns=["Phase", "Mean stride duration", 
                                                                     "standard deviation of stride duration",
                                                                     "Coefficient of variance of stride duration",
                                                                     "Number of strides","Duration"])
            
            
        
            self.exported_results_hand=pd.DataFrame(result_adj, columns=["Phase","Start Heel strike foot",
                                                                      "Stop Heel strike foot",
                                                                      "stride duration"])     
            
            
            
            self.lbl_mean_stridetime_HP.configure(text ='%.2f seconds'%(np.round(np.mean(result_adj[:,3]),decimals=5)))
            self.lbl_SDstride_HP.configure(text ='%.2f milliseconds'%(np.round(np.std(result_adj[:,3]),decimals=5)*1000))
            self.lbl_CVstride_HP.configure(text ='%.2f %%'%(np.round(np.std(result_adj[:,3])/np.mean(result_adj[:,3]),decimals=5)*100))
            self.lbl_Nstride_HP.configure(text ='%d'%(len(result_adj[:,3])))
            self.lbl_Nphase_HP.configure(text ='%d'%(n_phases))
            
            if self.right_excel or self.left_excel:
                self.Calculate_metrics(position="hand")
                
                hand_steps=self.True_positives_phases_hand.iloc[:,2]
                gaitup_steps=self.True_positives_phases_hand.iloc[:,1]
            
                self.CP_data_h.computeVarStride(fs=100,remove_outliers=False,N=3,use_smartstep=True,manual_peaks=hand_steps,use_peaks=True,
                                              pocket=False,remove_step=0)
            
                self.hand_strides=pd.DataFrame(self.CP_data_h.cycle_temp["detailed_stridetime"],columns=["Heel_strike","Foot_off","Stride_time"])

                self.CP_data_h.computeVarStride(fs=100,remove_outliers=False,N=3,use_smartstep=True,manual_peaks=gaitup_steps,use_peaks=True,
                                              pocket=False,remove_step=0)
            
                self.gaitup_hand_strides=pd.DataFrame(self.CP_data_h.cycle_temp["detailed_stridetime"],columns=["Heel_strike","Foot_off","Stride_time"])
                
                hand_index_todrop=self.hand_strides[(self.hand_strides.Stride_time > 3) | (self.hand_strides.Stride_time < 0) ].index
                
                self.hand_strides=self.hand_strides.drop(hand_index_todrop)
            
                self.gaitup_hand_strides=self.gaitup_hand_strides.drop(hand_index_todrop)
        
                gaitup_index_todrop=self.gaitup_hand_strides[(self.gaitup_hand_strides.Stride_time > 3) | (self.gaitup_hand_strides.Stride_time < 0) ].index
        
                self.gaitup_hand_strides=self.gaitup_hand_strides.drop(gaitup_index_todrop)
                
                self.hand_strides=self.hand_strides.drop(gaitup_index_todrop)

                
                
                
    def plot_graph(self,event=None,plot_name=""):
        try: 
            self.canvas.get_tk_widget().pack_forget()
            self.toolbar.destroy()
        except AttributeError: 
            pass     

        # Setup matplotlib canvas
        self.figure = Figure(figsize=(10, 9), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, master=self.fcontainer)
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        #set method on click
        self.cid = self.figure.canvas.mpl_connect('button_press_event', self.on_click_canvas)
        
        
        # Setup matplotlib toolbar (optional)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.fcontainer)
        self.toolbar.update()
        self.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        

        self.a = self.figure.add_subplot(511)
        self.a.plot(self.CP_data_phone.acc_magnitude,zorder=1)
    
        self.b = self.figure.add_subplot(512)
        self.b.plot(self.CP_data_phone.gyro_magnitude,zorder=1)
        
        if self.rightfoot or self.leftfoot:
            self.c = self.figure.add_subplot(513)
        if self.rightfoot:
            self.c.plot(self.acc_rf_magnitude,zorder=1)
        if self.leftfoot:
            self.c.plot(self.acc_lf_magnitude,zorder=1)
        
        if self.hand:
            self.d = self.figure.add_subplot(514)
            self.d.plot(self.acc_h_magnitude,zorder=1)
            self.e = self.figure.add_subplot(515)
            self.e.plot(self.gyro_h_magnitude,zorder=1)
            
        
        
        
        self.canvas.draw()
 
            
    def calculate_norm_accandgyro(self,gyro=[0],acc=[0]):
        
        matrix1=acc
        x=matrix1.iloc[:,0].values**2
        y=matrix1.iloc[:,1].values**2
        z=matrix1.iloc[:,2].values**2
        m=x+y+z
        mm=np.array([np.sqrt(i) for i in m])
        acc_magnitude=mm
        
        gyro_magnitude=0
        if len(gyro)==0:
            matrix2=gyro
            x=matrix2.iloc[:,0].values**2
            y=matrix2.iloc[:,1].values**2
            z=matrix2.iloc[:,2].values**2
            m=x+y+z
            mm=np.array([np.sqrt(i) for i in m])
            gyro_magnitude=mm
        
        return(acc_magnitude,gyro_magnitude)
    

    def Save_and_Export(self,event=None):

        # self.hand=False
        dfw1=self.exported_results_summary_phone
        dfw2=self.exported_results_phone
        
        
        
        phases=[]
        for line in self.treeview.get_children():
            phases.append(self.treeview.item(line)['values'])
        
        phases=np.vstack(phases)
        
        df_turns=pd.DataFrame(phases,columns=['Phase','Start','Stop','Steps_Turn'])
        
        
        filename=self.export_path+"\\new_exported_results.xlsx"
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        
        dfw1.to_excel(writer, sheet_name='Summary of results', startrow=1,index=False)
        
        dfw2.to_excel(writer, sheet_name='Strides and Heel strikes', startrow=1,
                      startcol=0,index=False)
        
        df_turns.to_excel(writer,sheet_name='Settings',
                          startrow=5,startcol=1,index=False)
        
        
        
        
        if self.right_excel or self.left_excel:
            dff1=self.exported_results_summary_feet
            dff1.to_excel(writer, sheet_name='Summary of results', startrow=len(dfw1)+4
                          ,index=False)
            dff2=self.exported_results_feet
            
            dff2.to_excel(writer, sheet_name='Strides and Heel strikes', startrow=1,
                          startcol=6,index=False)
            
            self.confusion_matrixs.to_excel(writer, sheet_name='Summary of results', startrow=1,startcol=len(dfw1.columns)+4,index=False)
        
            self.phone_strides.to_excel(writer, sheet_name='True positives2', startrow=1,
                          startcol=1,index=False)
            
            self.gaitup_strides.to_excel(writer, sheet_name='True positives2', startrow=1,
                          startcol=6,index=False)
            
            if self.hand:
                self.hand_strides.to_excel(writer, sheet_name='True positives2', startrow=1,
                          startcol=12,index=False)
                
                self.gaitup_hand_strides.to_excel(writer, sheet_name='True positives2', startrow=1,
                          startcol=18,index=False)
            
            
            
            
        if self.hand:
            dfh2=self.exported_results_hand
            dfh1=self.exported_results_summary_hand
            dfh1.to_excel(writer, sheet_name='Summary of results',
                          startrow=len(dfw1)+4+len(dff1)+3,index=False)
            
            dfh2.to_excel(writer, sheet_name='Strides and Heel strikes', startrow=1,
                          startcol=12,index=False)
            
            self.confusion_matrixs_hand.to_excel(writer, sheet_name='Summary of results', 
                                                 startrow=1+len(dfw1)+4+len(dff1)+2,
                                                 startcol=len(dfw1.columns)+4,index=False)
            

        worksheet1 = writer.sheets['Summary of results']
        worksheet2 = writer.sheets['Strides and Heel strikes']
        worksheet4=writer.sheets['True positives2']
        workbook  = writer.book
        
        worksheet3=writer.sheets['Settings']
        
        merge_format = workbook.add_format({'bold': True, 'font_color': 'red'})
        
        Miss_detected_format = workbook.add_format({'bg_color':   '#FFC7CE'})
        
        Miss_detected_format_hand = workbook.add_format({'bg_color':   '#FF7F00'})
        
        True_detected_format = workbook.add_format({'bg_color':   '#C6EFCE'})
        
        ############# worksheet 1###########
        worksheet1.merge_range(0,0,0,len(dfw1.columns)-1, 'Phone waist position', merge_format)
        if self.right_excel or self.left_excel:
            worksheet1.merge_range(0,len(dfw1.columns)+4,0,
                                   len(dfw1.columns)+4+len(self.confusion_matrixs.columns)-1, 
                                   'Confusion matrics step detection', merge_format)
            for i in self.start[0]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,1,i+2,1,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': True_detected_format})
    
            for i in self.start[1]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,7,i+2,7,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': Miss_detected_format})
                
            for i in self.stop[0]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,2,i+2,2,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': True_detected_format})
    
            for i in self.stop[1]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,8,i+2,8,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': Miss_detected_format})
                
            worksheet4.write_string(1,23,'RMSE of stride time')
            worksheet4.write_string(2,23,'Gaitup-phone mean stride time')
            worksheet4.write_string(3,23,'phone mean stride time')
            
            worksheet4.write_string(6,23,'Gaitup-phone SD stride time')
            worksheet4.write_string(7,23,'phone SD stride time')

             
            RMSE_pg=np.sqrt(np.mean((self.phone_strides["Stride_time"].values-self.gaitup_strides["Stride_time"].values)**2))*1000
            ST_g=np.mean(self.gaitup_strides["Stride_time"].values)
            ST_p=np.mean(self.phone_strides["Stride_time"].values)
            
            worksheet4.write_string(1,24,str(RMSE_pg))
            worksheet4.write_string(2,24,str(ST_g))
            worksheet4.write_string(3,24,str(ST_p))
            

            
            ST_g=np.std(self.gaitup_strides["Stride_time"].values)*1000
            ST_p=np.std(self.phone_strides["Stride_time"].values)*1000
            
            worksheet4.write_string(6,24,str(ST_g))
            worksheet4.write_string(7,24,str(ST_p))
            
            
            if self.hand:
                worksheet4.write_string(4,23,'Gaitup-hand mean stride time')
                worksheet4.write_string(5,23,'hand mean stride time')
                worksheet4.write_string(8,23,'Gaitup-hand SD stride time')
                worksheet4.write_string(9,23,'hand SD stride time')
                worksheet4.write_string(10,23,'RMSE of stride time hand gaitup')
                
                ST_gh=np.mean(self.gaitup_hand_strides["Stride_time"].values)
                ST_h=np.mean(self.hand_strides["Stride_time"].values)
                
                worksheet4.write_string(4,24,str(ST_gh))
                worksheet4.write_string(5,24,str(ST_h))
                
                ST_gh=np.std(self.gaitup_hand_strides["Stride_time"].values)*1000
                ST_h=np.std(self.hand_strides["Stride_time"].values)*1000
                
                worksheet4.write_string(8,24,str(ST_gh))
                worksheet4.write_string(9,24,str(ST_h))
            


            
                RMSE_hg=np.sqrt(np.mean((self.gaitup_hand_strides["Stride_time"].values-self.hand_strides["Stride_time"].values)**2))*1000
    
                worksheet4.write_string(10,24,str(RMSE_hg))
           
            
        
        
        worksheet1.write_string(1,len(dfw1.columns)+1,'Mean Stride time')
        worksheet1.write_string(1,len(dfw1.columns)+2,self.lbl_mean_stridetime.cget('text'))
        
        worksheet1.write_string(2,len(dfw1.columns)+1,'SD Stride time')
        worksheet1.write_string(2,len(dfw1.columns)+2,self.lbl_SDstride.cget('text'))
        
        worksheet1.write_string(3,len(dfw1.columns)+1,'Number of strides')
        worksheet1.write_string(3,len(dfw1.columns)+2,self.lbl_Nstride.cget('text'))
        
        worksheet1.write_string(4,len(dfw1.columns)+1,'Number of phases')
        worksheet1.write_string(4,len(dfw1.columns)+2,self.lbl_Nphase.cget('text'))

        ##############worksheet3 ########

        worksheet3.write_string(1,1,'Start_analysis_phone')
        worksheet3.write_string(1,2,self.ent_str_phone_waist.get())
        
        worksheet3.write_string(1,4,'Stop_analysis_phone')
        worksheet3.write_string(1,5,str(self.e_w))

        #######worksheet 2#########

        worksheet2.merge_range(0,0,0,4, 'Phone waist position', merge_format)
        
        # print(len(self.start[0]))
        # print(len(self.stop[0]))

        if self.right_excel or self.left_excel:
            
            worksheet1.merge_range(len(dfw1)+3,0,len(dfw1)+3,len(dfw1.columns)-1, 'feet gaitup position', merge_format)

            worksheet1.write_string(1+len(dfw1)+3,len(dfw1.columns)+1,'Mean Stride time')
            worksheet1.write_string(1+len(dfw1)+3,len(dfw1.columns)+2,self.lbl_mean_stridetime_GU.cget('text'))
        
            worksheet1.write_string(2+len(dfw1)+3,len(dfw1.columns)+1,'SD Stride time')
            worksheet1.write_string(2+len(dfw1)+3,len(dfw1.columns)+2,self.lbl_SDstride_GU.cget('text'))
        
            worksheet1.write_string(3+len(dfw1)+3,len(dfw1.columns)+1,'Number of strides')
            worksheet1.write_string(3+len(dfw1)+3,len(dfw1.columns)+2,self.lbl_Nstride_GU.cget('text'))
        
            worksheet1.write_string(4+len(dfw1)+3,len(dfw1.columns)+1,'Number of phases')
            worksheet1.write_string(4+len(dfw1)+3,len(dfw1.columns)+2,self.lbl_Nphase_GU.cget('text'))
                        
            ######
            worksheet3.write_string(2,1,'Start_analysis_Gaitup')
            worksheet3.write_string(2,2,self.ent_str_gaitup_feet.get())
            
            worksheet3.write_string(2,4,'Stop_analysis_Gaitup')
            worksheet3.write_string(2,5,str(self.e_g))
            ######
            
            worksheet2.merge_range(0,6,0,10, 'Feet Gaitup position', merge_format)
            
            if self.acquisition_stops:
                worksheet3.write_string(2,7,'NOTE: Acquisition of phone stops before end of test')
            
        if self.hand:
            worksheet1.merge_range(len(dfw1)+4+len(dff1)+2,
                                   0,len(dfw1)+4+len(dff1)+2,
                                   len(dfh1.columns)-1, 'Hand Gaitup position', merge_format)
            
            worksheet1.write_string(1+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+1,'Mean Stride time')
            worksheet1.write_string(1+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+2,self.lbl_mean_stridetime_HP.cget('text'))
            
            worksheet1.write_string(2+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+1,'SD Stride time')
            worksheet1.write_string(2+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+2,self.lbl_SDstride_HP.cget('text'))
            
            worksheet1.write_string(3+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+1,'Number of strides')
            worksheet1.write_string(3+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+2,self.lbl_Nstride_HP.cget('text'))
            
            worksheet1.write_string(4+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+1,'Number of phases')
            worksheet1.write_string(4+len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+2,self.lbl_Nphase_HP.cget('text'))
            
            worksheet1.merge_range(len(dfw1)+4+len(dff1)+2,len(dfw1.columns)+4,
                                   len(dfw1)+4+len(dff1)+2,
                                   len(dfw1.columns)+4+len(self.confusion_matrixs_hand.columns)-1, 
                                   'Confusion matrics step detection', merge_format)


            worksheet3.write_string(3,1,'Start_analysis_Hand')
            worksheet3.write_string(3,2,self.ent_str_gaitup_hand.get())
            
            worksheet2.merge_range(0,12,0,16, 'Hand Gaitup position', merge_format)
            
            for i in self.start_hand[0]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,13,i+2,13,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': True_detected_format})
    
            for i in self.start_hand[1]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,7,i+2,7,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': Miss_detected_format_hand})
                
            for i in self.stop_hand[0]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,14,i+2,14,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': True_detected_format})
    
            for i in self.stop_hand[1]:
                #(first_row, first_col, last_row, last_col)
                worksheet2.conditional_format(i+2,8,i+2,8,{'type': 'text',
                                           'criteria': 'containing',
                                           'value': '',
                                           'format': Miss_detected_format_hand})
            
        writer.save()
        
        print("saved")
    
    
    #callbacks for editable treeview
    def on_row_edit(self, event):
        # Get the column id and item id of the cell
        # that is going to be edited
        col, item = self.treeview.get_event_info()
        print(col)
        print(item)
        

        # Define the widget editor to be used to edit the column value
        if col in ('Start','Stop','Turning'):
            self.treeview.inplace_entry(col, item)

    def on_cell_changed(self, event):
        col, item = self.treeview.get_event_info()
        
        print('Column {0} of item {1} was changed'.format(col, item))

    def on_row_selected(self, event):
        print('Rows selected', event.widget.selection())
        
    def plot_stridetime(self):
        self.figure2 = Figure(figsize=(6, 5), dpi=100)
        self.canvas2 = FigureCanvasTkAgg(self.figure2, master=self.fcontainer2)
        self.canvas2.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas2, self.fcontainer2)
        self.toolbar.update()
        self.canvas2._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.b = self.figure2.add_subplot(111)
            
        self.b.plot(self.exported_results_phone['stride duration'],zorder=1)
        self.b.set_ylabel('Stride duration (s)')
        self.b.set_xlabel('Stride number')
        self.b.axhline(y=np.mean(self.exported_results_phone['stride duration']),label='Average Stride Time',color='r')
        self.b.axhline(y=np.mean(self.exported_results_phone['stride duration'])-np.std(self.exported_results_phone['stride duration']),linestyle='--',color='r',linewidth=0.5)
        self.b.axhline(y=np.mean(self.exported_results_phone['stride duration'])+np.std(self.exported_results_phone['stride duration']),linestyle='--',color='r',linewidth=0.5)
        self.b.text(0.8, 0.8,("Stride std:%s, Stride cov:%s"%(self.lbl_SDstride.cget("text"),self.lbl_CVstride.cget("text"))),size=10,transform=self.figure2.transFigure,ha="right", va="top", bbox=dict(facecolor='red', alpha=0.5))
        self.canvas2.draw()
        
        
    def calculate_DFA(self,data):
        """
        """
        strides_phone=[]
        strides_phone2=[]
        strides_gaitup=[]
        strides_gaitup2=[]
        strides_hand=[]
        strides_hand2=[]
        result_excel="_exported_results.xlsx"
        result_excel=os.path.join(self.path,result_excel)
        
        wb_result= load_workbook(result_excel)
        wb_result=wb_result["Strides and Heel strikes"]
        
        
        strides_phone=wb_result['D']
        strides_phone=strides_phone[2:]
        
        strides_phone2=[]
        for cl in strides_phone:
            if cl.value!=None:
                strides_phone2.append(cl.value)
            
        strides_phone2=np.hstack(strides_phone2).astype('float64')
        
        
        if self.right_excel or self.left_excel:          
            strides_gaitup=wb_result['J']
            strides_gaitup=strides_gaitup[2:]
            
            strides_gaitup2=[]
            for cl in strides_gaitup:
                if cl.value!=None:
                    strides_gaitup2.append(cl.value)
                
            strides_gaitup2=np.hstack(strides_gaitup2).astype('float64')
        
        if self.hand:
            strides_hand=wb_result['P']
            strides_hand=strides_hand[2:]
            
            
            strides_hand2=[]
            for cl in strides_hand:
                if cl.value!=None:
                    strides_hand2.append(cl.value)
                
            strides_hand2=np.hstack(strides_hand2).astype('float64')
        
        #read excel file
        data1=strides_phone2
        
        N=len(data1)
        f=2**(1/8)
        if N>1000:
            nvals=nld.logarithmic_n(16, N//9,f)
        else:
            nvals=nld.logarithmic_n(4, N//4,f)
            
        xx,yy=nld.dfa(data1, nvals=nvals, overlap=False, order=1, fit_trend="poly",fit_exp="poly", debug_plot=True, debug_data=True, plot_file=None)
        
        self.lbl_DFA_phone.configure(text ='%.2f '%(xx))

        if self.right_excel or self.left_excel:
            data2=strides_gaitup2
            
            N=len(data2)
            f=2**(1/8)
            if N>1000:
                nvals=nld.logarithmic_n(16, N//9,f)
            else:
                nvals=nld.logarithmic_n(4, N//4,f)
                
            xx,yy=nld.dfa(data2, nvals=nvals, overlap=False, order=1, fit_trend="poly",fit_exp="poly", debug_plot=True, debug_data=True, plot_file=None)
                    
            self.lbl_DFA_feet.configure(text ='%.2f '%(xx))

        if self.hand:
            data3=strides_hand2
            
            N=len(data3)
            f=2**(1/8)
            if N>1000:
                nvals=nld.logarithmic_n(16, N//9,f)
            else:
                nvals=nld.logarithmic_n(4, N//4,f)
                
            xx,yy=nld.dfa(data3, nvals=nvals, overlap=False, order=1, fit_trend="poly",fit_exp="poly", debug_plot=True, debug_data=True, plot_file=None)
    
            self.lbl_hand.configure(text ='%.2f '%(xx))


    def Calculate_lyapunov(self, event=None,tao=0,dim=0,three_dim=True):
        print("calculating lyapunov")
        if self.preprocessed==False or self.lyap_NBstrides != int((self.ent_Nbstridelyap.get())):
            self.lyap_NBstrides=int((self.ent_Nbstridelyap.get()))
            self.preprocessing_lyap(totalnumberofstride=self.lyap_NBstrides)

        tao=int((self.ent_embdim.get()))
        dim=int(self.ent_timedelay.get())
        
        print("calculating lyap on magnitude")
        _,self.debug=nld.lyap_r(self.acc_data, emb_dim=dim,lag=tao,
                                min_tsep=40, tau=0.01, trajectory_len=100*10, fit='poly', debug_plot=True,debug_data=True, plot_file="lyap", fit_offset=0)
    
        stride_duration=100
        plt.figure()
        d=self.debug[1]
        x=self.debug[0]/stride_duration
        plt.plot(x,d)
        
        j=stride_duration//2
        #plt.plot(debug[0],debug[1])
        z=np.polyfit(self.debug[0][0:j],self.debug[1][0:j],1)
        p = np.poly1d(z)
        plt.plot(self.debug[0][0:j]/stride_duration,p(self.debug[0][0:j]))
        lle=z[0]*100
        print("the lyap of acceleration norm is %.3f" %lle)
        self.lbl_lyapr.configure(text ='%.2f '%(lle))
        
        print("calculating lyap on three dimensions")
        _,self.debug=nld.lyap_r(self.threeD_acc_data[self.threeD_acc_data.columns[0]], 
                                emb_dim=dim,lag=tao, min_tsep=40, tau=0.01, trajectory_len=100*10, fit='poly', debug_plot=True,debug_data=True, plot_file="lyap", fit_offset=0)
        stride_duration=100
        plt.figure()
        d=self.debug[1]
        x=self.debug[0]/stride_duration
        plt.plot(x,d)
        
        j=stride_duration//2
        # plt.plot(debug[0],debug[1])
        z=np.polyfit(self.debug[0][0:j],self.debug[1][0:j],1)
        p = np.poly1d(z)
        plt.plot(self.debug[0][0:j]/stride_duration,p(self.debug[0][0:j]))
        lle=z[0]*100
        print("the lyap of X direction is %.3f" %lle)
        self.lbl_lyapx.configure(text ='%.2f '%(lle))
        
        _,self.debug=nld.lyap_r(self.threeD_acc_data[self.threeD_acc_data.columns[1]],
                                emb_dim=dim,lag=tao, min_tsep=40, tau=0.01, trajectory_len=100*10, fit='poly', debug_plot=True,debug_data=True, plot_file="lyap", fit_offset=0)
        stride_duration=100
        plt.figure()
        d=self.debug[1]
        x=self.debug[0]/stride_duration
        plt.plot(x,d)
        
        j=stride_duration//2
        #plt.plot(debug[0],debug[1])
        z=np.polyfit(self.debug[0][0:j],self.debug[1][0:j],1)
        p = np.poly1d(z)
        plt.plot(self.debug[0][0:j]/stride_duration,p(self.debug[0][0:j]))
        lle=z[0]*100
        print("the lyap of Y direction is %.3f" %lle)
        self.lbl_lyapy.configure(text ='%.2f '%(lle))
        
        _,self.debug=nld.lyap_r(self.threeD_acc_data[self.threeD_acc_data.columns[2]].values,
                                emb_dim=dim,lag=tao, min_tsep=40, tau=0.01, trajectory_len=100*10, fit='poly', debug_plot=True,debug_data=True, plot_file="lyap", fit_offset=0)
        stride_duration=100
        plt.figure()
        d=self.debug[1]
        x=self.debug[0]/stride_duration
        plt.plot(x,d)
        
        j=stride_duration//2
        #plt.plot(debug[0],debug[1])
        z=np.polyfit(self.debug[0][0:j],self.debug[1][0:j],1)
        p = np.poly1d(z)
        plt.plot(self.debug[0][0:j]/stride_duration,p(self.debug[0][0:j]))
        lle=z[0]*100
        print("the lyap of Z direction is %.3f" %lle)
        self.lbl_lyapz.configure(text ='%.2f '%(lle))

    def Calculate_embdimension(self,event=None,tao=10):
        print("calculating embedding dimension")
        if self.preprocessed==False or self.lyap_NBstrides != int((self.ent_Nbstridelyap.get())):
            self.lyap_NBstrides=int((self.ent_Nbstridelyap.get()))
            self.preprocessing_lyap(totalnumberofstride=self.lyap_NBstrides)
            
        self.f=dimension.fnn(self.acc_data, dim=[1,2,3,4,5,6,7,8,9,10], tau=tao, R=15.0, A=2.0, metric='euclidean', window=50,maxnum=None, parallel=True)
        plt.figure()
        plt.plot(self.f[0])
        plt.title("dim of magnitude")
        
        self.f=dimension.fnn(self.threeD_acc_data[self.threeD_acc_data.columns[0]], dim=[1,2,3,4,5,6,7,8,9,10], tau=tao, R=15.0, A=2.0, metric='euclidean', window=50,maxnum=None, parallel=True)
        plt.figure()
        plt.plot(self.f[0])
        plt.title("dim of accx")
        
        self.f=dimension.fnn(self.threeD_acc_data[self.threeD_acc_data.columns[1]], dim=[1,2,3,4,5,6,7,8,9,10], tau=tao, R=15.0, A=2.0, metric='euclidean', window=50,maxnum=None, parallel=True)
        plt.figure()
        plt.plot(self.f[0])
        plt.title("dim of accy")

        self.f=dimension.fnn(self.threeD_acc_data[self.threeD_acc_data.columns[2]], dim=[1,2,3,4,5,6,7,8,9,10], tau=tao, R=15.0, A=2.0, metric='euclidean', window=50,maxnum=None, parallel=True)
        plt.figure()
        plt.plot(self.f[0])
        plt.title("dim of accz")

    def preprocessing_lyap(self,totalnumberofstride=150):
        self.preprocessed=True
        
        acc,gyro=self.extract_signal_walking()
        
        gyro = pd.concat(gyro,ignore_index=True)
        acc = pd.concat(acc,ignore_index=True)
        index_step=[]

        result_excel="_exported_results.xlsx"
        result_excel=os.path.join(self.path,result_excel)
        wb_result= load_workbook(result_excel)
        wb_result=wb_result["Strides and Heel strikes"]
        
        index_foot1=wb_result['B']
        index_foot1=index_foot1[2:]
        
        # index_foot2=wb_result['C']
        # index_foot2=index_foot2[2:]
        
        index_step=[]
        for cl in index_foot1:
            if cl.value!=None:
                index_step.append(cl.value)
            
        index_step=np.hstack(index_step).astype('int')
        
        #calculating the norm of the signal when norms are removed and 
        #then normalising to a fixed number of strides and number of points
        self.CP_data_phone.calculate_norm_accandgyro(gyro=gyro,acc=acc)
        
        
        
        self.CP_data_phone.normal_signal(threeD_acc=acc,threeD_gyro=gyro,
                                         peaks=index_step,remove_outliers=False,N=3,
                                         Numberofstrides=totalnumberofstride,plot=False)
        
        print("creating acc_data")
        # retrieving the normalised signals
        self.acc_data=self.CP_data_phone.norma_acc_strides
        self.gyro_data=self.CP_data_phone.norma_acc_strides
        
        self.threeD_acc_data=self.CP_data_phone.norma_threeD_acc
        self.threeD_gyro_data=self.CP_data_phone.norma_threeD_gyro
        
    def Calculate_time_delay(self, event=None):
        
        print("calculating time delay")
        if self.preprocessed==False or self.lyap_NBstrides != int((self.ent_Nbstridelyap.get())):
            self.lyap_NBstrides=int((self.ent_Nbstridelyap.get()))
            self.preprocessing_lyap(totalnumberofstride=self.lyap_NBstrides)
            
        lag=delay.dmi(self.acc_data, maxtau=50, bins=64)
        plt.figure()
        plt.title("lag of magnitude")
        plt.plot(lag)
        
        lag=delay.dmi(self.threeD_acc_data[self.threeD_acc_data.columns[0]], maxtau=50, bins=64)
        plt.figure()
        plt.title("lag of x")
        plt.plot(lag)
        
        lag=delay.dmi(self.threeD_acc_data[self.threeD_acc_data.columns[1]], maxtau=50, bins=64)
        plt.figure()
        plt.title("lag of y")
        plt.plot(lag)
        
        lag=delay.dmi(self.threeD_acc_data[self.threeD_acc_data.columns[2]], maxtau=50, bins=64)
        plt.figure()
        plt.title("lag of z")
        plt.plot(lag)
        
    def calculate_entropy(self,event=None):
        
        print("calculating entropy")
        #Do the same normalisation as for lyapunov
        if self.preprocessed==False or self.lyap_NBstrides != int((self.ent_Nbstridelyap.get())):
            self.lyap_NBstrides=int((self.ent_Nbstridelyap.get()))
            self.preprocessing_lyap(totalnumberofstride=self.lyap_NBstrides)

        r=float((self.ent_tolerance.get()))
        m=int(self.ent_vect_length.get())
        r=None
        
        print("calculating sample entropy x")
        
        se_x=nld.sampen(self.threeD_acc_data[self.threeD_acc_data.columns[0]], emb_dim=m, tolerance=r,
                        debug_plot=False,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_x.configure(text ='%.2f '%(se_x))
        
        print(se_x)
        
        print("calculating sample entropy y")
        
        se_y=nld.sampen(self.threeD_acc_data[self.threeD_acc_data.columns[1]], emb_dim=m, tolerance=r,
                        debug_plot=False,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_y.configure(text ='%.2f '%(se_y))
        
        print(se_y)
        
        print("calculating sample entropy z")
        
        se_z=nld.sampen(self.threeD_acc_data[self.threeD_acc_data.columns[2]], emb_dim=m, tolerance=r,
                        debug_plot=False,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_z.configure(text ='%.2f '%(se_z))
        
        print(se_z)
        
        print("calculating sample entropy r")
        
        se_r=nld.sampen(self.acc_data, emb_dim=m, tolerance=r,
                        debug_plot=False,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_r.configure(text ='%.2f '%(se_r))
        
        print(se_r)
        
        rgc=None
        mgc=100
        
        print("calculating sample entropy gaitcycle x")
        
        se_x_gaitcycle=nld.sampen(self.threeD_acc_data[self.threeD_acc_data.columns[0]], 
                                  emb_dim=mgc,emb_dim_plus_one=mgc+100, tolerance=rgc,
                          debug_plot=False,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_gc_x.configure(text ='%.2f '%(se_x_gaitcycle))
        
        print(se_x_gaitcycle)
        
        print("calculating sample entropy gaitcycle y")
        
        se_y_gaitcycle=nld.sampen(self.threeD_acc_data[self.threeD_acc_data.columns[1]], 
                                  emb_dim=mgc,emb_dim_plus_one=mgc+100, tolerance=rgc,
                        debug_plot=True,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_gc_y.configure(text ='%.2f '%(se_y_gaitcycle))
        
        print(se_y_gaitcycle)
        
        print("calculating sample entropy gaitcycle z")
        
        se_z_gaitcycle=nld.sampen(self.threeD_acc_data[self.threeD_acc_data.columns[2]], 
                                  emb_dim=mgc,emb_dim_plus_one=mgc+100, tolerance=rgc,
                        debug_plot=False,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_gc_z.configure(text ='%.2f '%(se_z_gaitcycle))
        
        print(se_z_gaitcycle)
        
        print("calculating sample entropy gaitcycle r")
        
        se_r_gaitcycle=nld.sampen(self.acc_data, emb_dim=mgc,emb_dim_plus_one=mgc+100, tolerance=rgc,
                        debug_plot=False,
                        debug_data=False, plot_file=None)
        
        self.lbl_se_gc_r.configure(text ='%.2f '%(se_r_gaitcycle))
        
        print(se_r_gaitcycle)
        
        ##reading stride times from excel file
        strides_phone=[]
        strides_phone2=[]
        strides_gaitup=[]
        strides_gaitup2=[]
        strides_hand=[]
        strides_hand2=[]
        result_excel="_exported_results.xlsx"
        result_excel=os.path.join(self.path,result_excel)
        
        wb_result= load_workbook(result_excel)
        wb_result=wb_result["Strides and Heel strikes"]
        
        
        strides_phone=wb_result['D']
        strides_phone=strides_phone[2:]
        
        strides_phone2=[]
        for cl in strides_phone:
            if cl.value!=None:
                strides_phone2.append(cl.value)
            
        strides_phone2=np.hstack(strides_phone2).astype('float64')
        
        
        if self.right_excel or self.left_excel:          
            strides_gaitup=wb_result['J']
            strides_gaitup=strides_gaitup[2:]
            
            strides_gaitup2=[]
            for cl in strides_gaitup:
                if cl.value!=None:
                    strides_gaitup2.append(cl.value)
                
            strides_gaitup2=np.hstack(strides_gaitup2).astype('float64')
        
        if self.hand:
            strides_hand=wb_result['P']
            strides_hand=strides_hand[2:]
            
            
            strides_hand2=[]
            for cl in strides_hand:
                if cl.value!=None:
                    strides_hand2.append(cl.value)
                
            strides_hand2=np.hstack(strides_hand2).astype('float64')
        
        #read excel file
        data1=strides_phone2
        
        se_stridetime_phone=nld.sampen(data1, emb_dim=m, tolerance=None,debug_plot=False,
                        debug_data=False, plot_file=None)

        self.lbl_se_stridetime.configure(text ='%.2f '%(se_stridetime_phone))

        # if self.right_excel or self.left_excel:
        #     data2=strides_gaitup2
            

        #     self.lbl_DFA_feet.configure(text ='%.2f '%(xx))

        # if self.hand:
        #     data3=strides_hand2
            

        #     self.lbl_hand.configure(text ='%.2f '%(xx))
        
        
        

        
        
        
        
        
        
        
        
        
        

                
    def clear_graph(self,event=None):

        self.canvas.get_tk_widget().pack_forget() 
        self.toolbar.destroy()
        
        
    def quit(self, event=None):
        self.mainwindow.quit()
        
        

    def run(self):
        self.mainwindow.mainloop()
        
        
        
if __name__ == '__main__':
    app = MyApplication()
    app.run()