# -*- coding: utf-8 -*-
"""
Created on Thu Jan 27 15:43:42 2022

@author: al-abiad
"""
import os
import tkinter as tk
import pygubu
from CPclass import phone as CP
import numpy as np
from matplotlib.backends.backend_tkagg import ( FigureCanvasTkAgg, NavigationToolbar2Tk)# Implement the default Matplotlib key bindings.
from matplotlib.figure import Figure
import pandas as pd
from os import walk
from scipy.signal import find_peaks
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from scipy.signal import find_peaks

CURRENT_DIR = os.path.abspath(os.path.dirname(__file__))
pickable_artists = []

class Label_data_app(object):
    
    def __init__(self):
        
        self.builder = builder = pygubu.Builder()
        
        builder.add_from_file(os.path.join(CURRENT_DIR, 'Label_app.ui'))
        
        self.mainwindow = builder.get_object('mainwindow')
        
        self.fcontainer = builder.get_object('fcontainer')
        
        self.filepath = builder.get_object('filepath')
        
        self.treeview=self.builder.get_object('treeview_peaks')
        
        self.btn_detect_step=self.builder.get_object('btn_detect_peaks')
        
        self.btn_save=self.builder.get_object('btn_save')
        
        self.speed = self.builder.get_object('speed')
        
        builder.connect_callbacks(self)
        
        
    def on_path_changed(self, event=None):
        # Get the path choosed by the user
        print("you are in on path changed")

        
        self.path = self.filepath.cget('path')
        self.export_path=self.path
        
        phone_dir=os.path.join(self.path,"telephone" )
        for root, dirs, files in os.walk(phone_dir):
            for file in files:
                if file.endswith(".txt"):
                    self.phone_direct=root
                break
        #phone directory is self.phone_direct
        print(self.phone_direct)

        # show paths in label
        _, _, self.files = next(walk(self.phone_direct))    
    

    def Detect_steps(self,event=None):
        print("detecting steps")
        
        self.read_gaitup_phone_csv()
        #----DETECTION OF STEPS

        self.align_signals()
        
        self.allsteps=[]
        
        #detect steps
        
        for i in range(0,len(self.start_turn)):
            acc_mag=self.acc_w_magnitude[self.start_turn[i]:self.stop_turn[i]]
            peak_index,_= find_peaks(acc_mag,distance=30,height=(None,None),
                                     prominence=(None,None),width=(None,None),
                                     threshold=(None,None)) # hand low speed prominence 1
            
            
            peak_index=peak_index+self.start_turn[i]
            
            self.allsteps.extend(peak_index)
        
        for p in self.allsteps:
            pt = self.b.scatter(p, 10,color='k',marker="|",s=10000000000,zorder=2,picker=5)
            pickable_artists.append(pt)

        
        
        



        if (len(self.treeview.get_children())!=0):
            self.treeview.delete(*self.treeview.get_children())
           
        for d in self.allsteps:
            self.treeview.insert('', tk.END, values=d)
            

            
        # for p in self.allsteps:
        #     pt = self.a.scatter(p, 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
        #     pickable_artists.append(pt)
            
        # self.canvas.draw()
        

        
        
    def plot_and_calculate_start(self,event=None):
        print("doing") 
        self.read_gaitup_phone_csv()
        
    def read_gaitup_phone_csv(self):
        """
        read the csv files of gaitup: left foot, right foot, hand into a dataframe
        interpolates and transform into CP object, to filter data from hand
        interpolate to 100 Hz 

        Returns
        -------
        None.

        """
        gaitup_dir=os.path.join(self.path,"Gaitup" )
        self.hand=True
        self.leftfoot=True
        self.rightfoot=True
        #hand norm
        filename=os.path.join(gaitup_dir,"hand.csv" )
        col_list = ["Time", "Gyro X","Gyro Y","Gyro Z","Accel X", "Accel Y","Accel Z"]
        try:
            df_h=pd.read_csv(filename, delimiter=",",skiprows=[0],usecols=col_list)

        except:
            self.hand=False
            print("no hand file")
            

            
        if self.hand:
            df_h = df_h.iloc[1:]
            df_h=df_h.astype('float32')
            
            self.df_h=df_h
            time=self.df_h['Time'].values
            
            #--interpolate manually--#
            self.df_h=self.df_h.set_index('Time')
            fs=100
            t_t=np.linspace(0, time[-1], num=np.int(time[-1]*fs), 
                            endpoint=True,dtype=np.float32)
            self.df_h=self.df_h.reindex(self.df_h.index.union(t_t))
            self.df_h=self.df_h.interpolate(method='linear', limit_direction='both',
                                            axis=0)
            self.df_h=self.df_h[self.df_h.index.isin(pd.Index(t_t))]
            self.df_h.index=np.around(self.df_h.index.astype('float32'),decimals=4)
            time =self.df_h.index
            acc_h=self.df_h.iloc[:,3:6]
            acc_h=acc_h.mul(9.8)
            gyro_h=self.df_h.iloc[:,0:3]
            
            
            #--transform into CP-object--#
            self.CP_data_h=CP(acc=acc_h,gyro=gyro_h,app="manual_entry")
            self.CP_data_h.deg2rad()
            self.CP_data_h.filter_data(acc=self.CP_data_h.acc_interp,gyro=self.CP_data_h.gyro_interp,N=10,fc=2,fs=100)
            self.CP_data_h.calculate_norm_accandgyro(gyro=self.CP_data_h.gyro_filtered,acc=self.CP_data_h.acc_filtered)
            self.acc_h_magnitude=self.CP_data_h.acc_magnitude
            self.gyro_h_magnitude=self.CP_data_h.gyro_magnitude
            
            #--Detect start of walk--#
            s_h=self.CP_data_h.detectstartofwalk(sig1=self.CP_data_h.acc_magnitude,thresh=3)

            self.ent_str_gaitup_hand.delete(0,tk.END)
            self.ent_str_gaitup_hand.insert(0,s_h)
        
        #--read left foot csv--#
        filename=os.path.join(gaitup_dir,"left_foot.csv" ) 
        
        try:
            col_list = ["Time", "Gyro X","Gyro Y","Gyro Z","Accel X", "Accel Y","Accel Z"]
            df_lf=pd.read_csv(filename, delimiter=",",skiprows=[0],usecols=col_list)

        except:
            self.leftfoot=False
            print("no leftfoot file")
        
        if self.leftfoot:
            df_lf = df_lf.iloc[1:]
            df_lf=df_lf.astype('float32')
            
            self.df_lf=df_lf
            time=self.df_lf['Time'].values
            #interpolate and filter
            self.df_lf=self.df_lf.set_index('Time')
            fs=100
            t_t=np.linspace(0, time[-1], num=np.int(time[-1]*fs), 
                            endpoint=True,dtype=np.float32)
    
            self.df_lf=self.df_lf.reindex(self.df_lf.index.union(t_t))
            self.df_lf=self.df_lf.interpolate(method='linear', limit_direction='both',
                                            axis=0)
            self.df_lf=self.df_lf[self.df_lf.index.isin(pd.Index(t_t))]
    
            self.df_lf.index=np.around(self.df_lf.index.astype('float32'),decimals=4)
    
            time =self.df_lf.index
            
            acc_lf=self.df_lf.iloc[:,3:6]
            gyro_lf=self.df_lf.iloc[:,0:3]
            
            #--create cp object from csv of foot--#
            self.CP_data_lf=CP(acc=acc_lf,gyro=gyro_lf,app="manual_entry")
            self.CP_data_lf.calculate_norm_accandgyro(gyro=self.CP_data_lf.gyro_interp,acc=self.CP_data_lf.acc_interp)
            self.acc_lf_magnitude=self.CP_data_lf.acc_magnitude
            self.gyro_lf_magnitude=self.CP_data_lf.gyro_magnitude
            s_lf=self.CP_data_lf.detectstartofwalk(sig1=self.CP_data_lf.acc_magnitude,thresh=5)
            s_f=s_lf
        #--read right foot csv--#
        filename=os.path.join(gaitup_dir,"right_foot.csv" )
        
        try:
            col_list = ["Time", "Gyro X","Gyro Y","Gyro Z","Accel X", "Accel Y","Accel Z"]
            df_rf=pd.read_csv(filename, delimiter=",",skiprows=[0],usecols=col_list)

        except:
            self.rightfoot=False
            print("no rightfoot file")
        
        if self.rightfoot:
            #right foot
            df_rf = df_rf.iloc[1:]
            df_rf=df_rf.astype('float32')
            
            self.df_rf=df_rf
            time=self.df_rf['Time'].values
            #interpolate and filter
            self.df_rf=self.df_rf.set_index('Time')
            fs=100
            t_t=np.linspace(0, time[-1], num=np.int(time[-1]*fs), 
                            endpoint=True,dtype=np.float32)
    
            self.df_rf=self.df_rf.reindex(self.df_rf.index.union(t_t))
            self.df_rf=self.df_rf.interpolate(method='linear', limit_direction='both',
                                            axis=0)
            self.df_rf=self.df_rf[self.df_rf.index.isin(pd.Index(t_t))]
    
            self.df_rf.index=np.around(self.df_rf.index.astype('float32'),decimals=4)
    
            time =self.df_rf.index
            
            acc_rf=self.df_rf.iloc[:,3:6]
            gyro_rf=self.df_rf.iloc[:,0:3]
            # read right into cp object
            self.CP_data_rf=CP(acc=acc_rf,gyro=gyro_rf,app="manual_entry")
            self.CP_data_rf.calculate_norm_accandgyro(gyro=self.CP_data_rf.gyro_interp,acc=self.CP_data_rf.acc_interp)
            self.acc_rf_magnitude=self.CP_data_rf.acc_magnitude
            self.gyro_rf_magnitude=self.CP_data_rf.gyro_magnitude
            s_rf=self.CP_data_rf.detectstartofwalk(sig1=self.CP_data_rf.acc_magnitude,thresh=5)
            s_f=s_rf
            
        if self.leftfoot and self.rightfoot:    
            #start of walking in foot signal is the minimum from right and left
            s_f=np.minimum(s_rf,s_lf)

            
        
        
        
        #read phone data and detect start#
        self.CP_data_phone=CP(self.phone_direct,app="geoloc")
        self.CP_data_phone.interpolategyrnacc(fs=100)
        self.CP_data_phone.filter_data(acc=self.CP_data_phone.acc_interp,gyro=self.CP_data_phone.gyro_interp,N=10,fc=3,fs=100)
        self.CP_data_phone.calculate_norm_accandgyro(gyro=self.CP_data_phone.gyro_filtered,acc=self.CP_data_phone.acc_filtered)
        self.acc_w_magnitude=self.CP_data_phone.acc_magnitude
        
        self.s_w=self.CP_data_phone.detectstartofwalk(sig1=self.CP_data_phone.acc_magnitude,thresh=2)

        self.e_w=int(len(self.CP_data_phone.acc_magnitude)-self.CP_data_phone.detectstartofwalk(sig1=self.CP_data_phone.acc_magnitude[::-1]))
        print("last walk inverse")
        print(self.CP_data_phone.detectstartofwalk(sig1=self.CP_data_phone.acc_magnitude[::-1]))
        
        print("length of signal")
        print(len(self.CP_data_phone.acc_magnitude))
        print("end of walk")
        print(self.e_w)

    def read_gaitup_excel(self,event=None):
        self.right_excel=False
        self.left_excel=False
        gaitup_dir=os.path.join(self.path,"Gaitup_turn" )
        for root, dirs, files in os.walk(gaitup_dir):
            for file in files:
                if file.endswith(".xlsx"):
                    if 'right' in file:
                        right_step_file= file
                        self.right_excel=True
                        print(file)
                    if 'left' in file:
                        left_step_file= file
                        self.left_excel=True
                        print(file)
        if self.right_excel:
            #right 
            right_file=os.path.join(gaitup_dir,right_step_file)
            
            
            wb_right= load_workbook(right_file)
            ws_right=wb_right["Sheet1"]
            
            start_analysis_gaitup_right = ws_right['B14'].value
            stop_analysis_gaitup_right=ws_right['D14'].value
            
            ColD_right=ws_right['D']
            ColD_right=ColD_right[25::]
            
            turns_HS_right=[]
            straightwalk_HS_right=[]
            i=0
            for cl in ColD_right:
                if cl.value!= None:
                    i=i+1
                    if cl.font.color != None and type(cl.font.color.rgb) == str:
                        x=cl.value+start_analysis_gaitup_right
                        turns_HS_right.append([i,x])
                    else:
                        x=cl.value+start_analysis_gaitup_right
                        straightwalk_HS_right.append([i,x])
                    
            turns_HS_right=np.vstack(turns_HS_right)
            straightwalk_HS_right=np.vstack(straightwalk_HS_right)
            
            #get index of straight walking steps in the main csv signal
            time=self.CP_data_rf.acc_interp.index.values
            straightwalk_HS_right_index=[]
            for hs in straightwalk_HS_right[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    straightwalk_HS_right_index.append(t1)
                else:
                    straightwalk_HS_right_index.append(t2)
            
            #get index of turn walking steps in the main csv signal
            Turn_HS_right_index=[]
            for hs in turns_HS_right[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    Turn_HS_right_index.append(t1)
                else:
                    Turn_HS_right_index.append(t2)
                    
            self.HS_r_index=np.array(straightwalk_HS_right_index)
            self.start_excel_right=start_analysis_gaitup_right
                    
        if self.left_excel:
                
            #left
            left_file=os.path.join(gaitup_dir,left_step_file)
            
            
            wb_left= load_workbook(left_file)
            ws_left=wb_left["Sheet1"]
            
            start_analysis_gaitup_left = ws_left['B14'].value
            stop_analysis_gaitup_left=ws_left['D14'].value
            
            ColD_left=ws_left['D']
            ColD_left=ColD_left[25::]
            
            turns_HS_left=[]
            straightwalk_HS_left=[]
            i=0
            for cl in ColD_left:
                if cl.value!= None:
                    i=i+1
                    if cl.font.color != None and type(cl.font.color.rgb) == str:
                        x=cl.value+start_analysis_gaitup_left
                        turns_HS_left.append([i,x])
                    else:
                        x=cl.value+start_analysis_gaitup_left
                        straightwalk_HS_left.append([i,x])
                    
            turns_HS_left=np.vstack(turns_HS_left)
            straightwalk_HS_left=np.vstack(straightwalk_HS_left)
            
            #get index of straight walking steps in the main csv signal
            time=self.CP_data_lf.acc_interp.index.values
            straightwalk_HS_left_index=[]
            for hs in straightwalk_HS_left[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    straightwalk_HS_left_index.append(t1)
                else:
                    straightwalk_HS_left_index.append(t2)
            
            #get index of turn walking steps in the main csv signal
            Turn_HS_left_index=[]
            for hs in turns_HS_left[:,1]:
                t1=np.where(time>hs)[0][0]
                t2=t1-1
                if np.abs(hs-time[t1])<=np.abs(hs-time[t2]):
                    Turn_HS_left_index.append(t1)
                else:
                    Turn_HS_left_index.append(t2)
                    
            self.HS_l_index=np.array(straightwalk_HS_left_index)
            self.start_excel_left=start_analysis_gaitup_left   
        # straight_walking_periods_left=self.ranges(straightwalk_HS_left[:,0])
        
        # turning_periods_left=self.ranges(turns_HS_left[:,0],d=2)
        
        if self.left_excel and self.right_excel:
            turn_right_left=np.concatenate((turns_HS_left[:,1],turns_HS_right[:,1]))
        
        elif self.left_excel:
            turn_right_left=turns_HS_left[:,1]
        
        elif self.right_excel:
            turn_right_left=turns_HS_right[:,1]
            
        if self.left_excel or self.right_excel:
            turn_right_left=turn_right_left*100
            turn_right_left=turn_right_left.astype("int")
            turn_right_left.sort()
            turn_right_left=np.unique(turn_right_left)
            
            self.turning_periods_left_right=self.ranges(turn_right_left,d=400)
            self.turning_periods_left_right=np.array(self.turning_periods_left_right)
            self.turning_periods_left_right=np.vstack(self.turning_periods_left_right)
            print(self.turning_periods_left_right)
        
        #removing dupliactes in case turn is last step

            i=0
            while i<len(self.turning_periods_left_right):
                if self.turning_periods_left_right[i,0]==self.turning_periods_left_right[i,1]:
                    self.turning_periods_left_right=np.delete(self.turning_periods_left_right,i,0)
                i=i+1
    
            print(self.turning_periods_left_right)
            
    def crop_signals(self,event=None,end="shortest"):
        
        
        try:
            f = open(self.path+"\\synch.txt", "r")
            n=f.readline().split(",")
            print(n)
            if n[0]!='':
                s_f=int(n[0])

                
            if n[1]!='':
                s_h=int(n[1])    

            if n[2]!='':
                s_w=int(n[2])

                stop_phone=self.e_w-s_w
            print("start and stop taken from text file")
        except:
            print("text file of synch doesnt exist")
        


        stop_ind_feet=stop_phone

        
        if self.rightfoot:
            self.CP_data_rf.manual_crop(ind_start=s_f)
            
        if self.leftfoot:
            self.CP_data_lf.manual_crop(ind_start=s_f)
            
        if self.left_excel or self.right_excel:
            print("aligning turns start")
            print(s_f)
            self.turning_periods_left_right=self.turning_periods_left_right-s_f
        
        if self.right_excel:
            self.HS_r_index=self.HS_r_index-s_f
            stop_ind_feet=self.HS_r_index[-1]
            self.e_g=stop_ind_feet
            
        if self.left_excel:
            self.HS_l_index=self.HS_l_index-s_f
            stop_ind_feet=self.HS_l_index[-1]
            self.e_g=stop_ind_feet
            
        if self.left_excel and self.right_excel:
            stop_ind_feet=np.maximum(self.HS_l_index[-1],self.HS_r_index[-1])+1
            self.e_g=stop_ind_feet

            
        self.CP_data_phone.manual_crop(ind_start=s_w)
        if self.hand:
            print("hand is available")
            self.CP_data_h.manual_crop(ind_start=s_h)
            
        self.acquisition_stops=False
        if end=="shortest":
            print("shortest")
            print(stop_ind_feet)
            if self.left_excel or self.right_excel:
                if stop_phone<stop_ind_feet-1000:
                    print("the phone stops acquisition before end of trial")
                    self.acquisition_stops=True
                stop_ind_feet=np.minimum(stop_phone,stop_ind_feet)
                print(stop_ind_feet)
                
                ind_turn=np.where((self.turning_periods_left_right[:,0]<=stop_ind_feet)&(self.turning_periods_left_right[:,1]<=stop_ind_feet))[0]
                self.turning_periods_left_right=self.turning_periods_left_right[ind_turn,:]
                
                if self.right_excel:
                    ind_foot=np.where(self.HS_r_index<=stop_ind_feet)[0]
                    self.HS_r_index=self.HS_r_index[ind_foot]
                if self.left_excel:
                    ind_foot=np.where(self.HS_l_index<=stop_ind_feet)[0]
                    self.HS_l_index=self.HS_l_index[ind_foot]

        #manual crop the interp#
        if self.rightfoot:
            self.CP_data_rf.manual_crop(ind_start=0,ind_stop=stop_ind_feet)
        if self.leftfoot:
            self.CP_data_lf.manual_crop(ind_start=0,ind_stop=stop_ind_feet)
        self.CP_data_phone.manual_crop(ind_start=0,ind_stop=stop_ind_feet)
        if self.hand:
            self.CP_data_h.manual_crop(ind_start=0,ind_stop=stop_ind_feet) 
            
    def recalculate_filter_norm(self):
        
        if self.leftfoot:
            self.CP_data_lf.calculate_norm_accandgyro(gyro=self.CP_data_lf.gyro_interp,
                                                  acc=self.CP_data_lf.acc_interp)
            self.acc_lf_magnitude=self.CP_data_lf.acc_magnitude
        if self.rightfoot:
            self.CP_data_rf.calculate_norm_accandgyro(gyro=self.CP_data_rf.gyro_interp,
                                                  acc=self.CP_data_rf.acc_interp)
            self.acc_rf_magnitude=self.CP_data_rf.acc_magnitude
        
        if self.hand:
            self.CP_data_h.filter_data(acc=self.CP_data_h.acc_interp,
                                       gyro=self.CP_data_h.gyro_interp)
            self.CP_data_h.calculate_norm_accandgyro(gyro=self.CP_data_h.gyro_filtered,
                                                     acc=self.CP_data_h.acc_filtered)
            self.acc_h_magnitude=self.CP_data_h.acc_magnitude
            self.gyro_h_magnitude=self.CP_data_h.gyro_magnitude

        self.CP_data_phone.filter_data(acc=self.CP_data_phone.acc_interp,
                                       gyro=self.CP_data_phone.gyro_interp)
        self.CP_data_phone.calculate_norm_accandgyro(gyro=self.CP_data_phone.gyro_filtered,
                                                     acc=self.CP_data_phone.acc_filtered)
        self.acc_w_magnitude=self.CP_data_phone.acc_magnitude
        self.gyro_w_magnitude=self.CP_data_phone.gyro_magnitude
    
    def align_signals(self):
        print("aligning")
        self.read_gaitup_excel()
        self.crop_signals()
        self.recalculate_filter_norm()
        self.plot_aligned_signals()
        self.Detect_turns()

        
        
    def plot_aligned_signals(self):
        
        try: 
            self.canvas4.get_tk_widget().pack_forget()
            self.toolbar.destroy()
        except AttributeError: 
            pass 

        self.figure4 = Figure(figsize=(10, 9), dpi=100)
        self.canvas4 = FigureCanvasTkAgg(self.figure4, master=self.fcontainer)
        self.canvas4.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas4, self.fcontainer)
        self.toolbar.update()
        self.canvas4._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        
        self.cid = self.figure4.canvas.mpl_connect('button_press_event', self.on_click_canvas)
        
        color = 'tab:blue'
        self.b = self.figure4.add_subplot(111)
        self.b.plot(self.acc_w_magnitude,zorder=1)
        self.b.plot(self.gyro_w_magnitude,zorder=1)
        
        if self.hand:
            self.b.plot(self.acc_h_magnitude,zorder=1)
            self.b.plot(self.gyro_h_magnitude,zorder=1)
            
            
        # if self.leftfoot:    
        #     self.b.plot(self.acc_lf_magnitude,zorder=1)
        # if self.rightfoot:
        #     self.b.plot(self.acc_rf_magnitude,zorder=1)
        
        self.b.set_ylabel('Acc m/s\u00b2', color=color)
        self.b.title.set_text("allsignals")
        self.b.set_xticks([])
        
        self.canvas4.draw()
        
        

    def calculate_norm_accandgyro(self,gyro=[0],acc=[0]):
        
        matrix1=acc
        x=matrix1.iloc[:,0].values**2
        y=matrix1.iloc[:,1].values**2
        z=matrix1.iloc[:,2].values**2
        m=x+y+z
        mm=np.array([np.sqrt(i) for i in m])
        acc_magnitude=mm
        
        gyro_magnitude=0
        if len(gyro)==0:
            matrix2=gyro
            x=matrix2.iloc[:,0].values**2
            y=matrix2.iloc[:,1].values**2
            z=matrix2.iloc[:,2].values**2
            m=x+y+z
            mm=np.array([np.sqrt(i) for i in m])
            gyro_magnitude=mm
        
        return(acc_magnitude,gyro_magnitude)
    
    
    def on_click_canvas(self,event):

        print('inaxes:', event.inaxes)
        
        if self.toolbar.mode =="":
            print("toolbar is not selected")
            
            if event.inaxes is not None and not hasattr(event, 'already_picked'):
                
                remove = [artist for artist in pickable_artists if artist.contains(event)[0]]
                print("--remove--")
                print(remove)
                print("---pickable_artists---")
                print(pickable_artists)
                x=event.xdata.astype('int')
                
                if not remove:
                    # add a pt
                    pt = self.b.scatter(x, 10,color='k',marker="|",s=10000000000,zorder=2,picker=5)
                    pickable_artists.append(pt)
                    self.allsteps=np.append(self.allsteps,x)
                    print("append")
                    self.refresh_treeview()

                else:
                    
                    self.allsteps=[i for i in self.allsteps if i>x+10 or i<x-10]
                    self.refresh_treeview()
                    
                    pickable_artists.remove(remove[0])
                    for artist in remove:
                        artist.remove()
                
                self.canvas4.draw()
            else:
                print ('Clicked ouside axes bounds but inside plot window')
        else:
            print("toolbar is selected")
            
            
    def refresh_treeview(self):
        
        self.allsteps=np.sort(self.allsteps)

        if (len(self.treeview.get_children())!=0):
            self.treeview.delete(*self.treeview.get_children())
            
        for d in self.allsteps:
            self.treeview.insert('', tk.END, values=d)

        
    def clear_graph(self,event=None):

        self.canvas4.get_tk_widget().pack_forget() 
        self.toolbar.destroy()
    
 
    def Detect_turns(self,method="gyro_peak",event=None):
        if self.left_excel or self.right_excel:
            method="extract gaitup"
        else:
            method=="gyro_peak"
        if method=="gyro_peak":
            self.phase=[]
            str_walk=0
            
            s_w=int((self.ent_str_phone_waist.get()))
            
            stop_phone=int((self.ent_stp_walk.get()))-s_w
            stp_walk=stop_phone
            
            turn=int(self.ent_turn.get())
            
            self.CP_data_phone.peakdet_m2(acc=False,plot_peak=True,detect_turn=True)
            self.peaks=self.CP_data_phone.peakandvalley["peak_index"]
    
            self.peaks=self.peaks[self.peaks>str_walk]
            self.peaks=self.peaks[self.peaks<stp_walk]
            
            self.peaks2 = [str_walk, *self.peaks, stp_walk]
    
            for i in range(0,len(self.peaks2)-1):
                self.phase.append((i+1,self.peaks2[i],self.peaks2[i+1],turn))
             
            self.peaks2=np.array(self.peaks2)
            
            if (len(self.treeview.get_children())!=0):
                self.treeview.delete(*self.treeview.get_children())
                
            for d in self.phase:
                self.treeview.insert('', tk.END, values=d)
    
            self.btn_detect_step["state"] = "normal"
            
            print(len(self.treeview.get_children()))
            
            for p in self.phase:
                pt = self.b.scatter(p[1], 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
                pt = self.b.scatter(p[2], 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)
            
            self.turn_peak_ind=self.peaks2
            
            self.canvas4.draw()
        elif method=="extract gaitup":
            print("extracting from gaitup")
            start_turn=self.turning_periods_left_right[:,0]
            stop_turn=self.turning_periods_left_right[:,1]
            
            self.start_turn=start_turn
            self.stop_turn=stop_turn
            
            self.phase=[]
            
            self.phase.append((0,0,start_turn[0],0))
            
            for i in range(0,len(start_turn)-1):
                if stop_turn[i]==start_turn[i]:
                    self.phase.append((i+1,stop_turn[i]+100,start_turn[i+1],0))
                else:   
                    self.phase.append((i+1,stop_turn[i],start_turn[i+1],0))
                     
            self.phase.append((i+2,stop_turn[-1],len(self.CP_data_phone.acc_magnitude),0))
            
            if (len(self.treeview.get_children())!=0):
                self.treeview.delete(*self.treeview.get_children())
                
            for d in self.phase:
                self.treeview.insert('', tk.END, values=d)
                
            for p in self.phase:
                print("in plot")

                #stop turn
                pt2 = self.b.scatter(p[1], 10,color='g',marker="|",s=10000000000,zorder=2,picker=5)
                #start turn
                pt1 = self.b.scatter(p[2], 10,color='red',marker="|",s=10000000000,zorder=2,picker=5)

                
                
            self.btn_detect_step["state"] = "normal"
            phases=np.vstack(self.phase)
            self.canvas4.draw()
            self.turn_stop_ind=phases[:,1]
            self.turn_start_ind=phases[:,2]

            print("gyro_turn_period_hilbert")

            #########################################################
            turns=self.window_rms(self.CP_data_phone.gyro_magnitude, window_size=300)
            peak_index,_= find_peaks(turns,distance=500)

            gyro_max=np.mean(turns[peak_index])
            
            # gyro_max=np.mean(turns[str_walk:stp_walk])
            print(gyro_max)
            gyro_binary=np.where(turns>gyro_max)[0]
            ranges=self.ranges(gyro_binary)
            turn_array=np.zeros(len(self.CP_data_phone.gyro_magnitude))

            for (start,end) in ranges:
                if start+50<end:
                    turn_array[start:end]=1
            
            #########################################################

            
            plt.figure()
            plt.plot(self.CP_data_phone.gyro_magnitude)
            plt.plot(turn_array)
            plt.plot(turns)
            
            self.turns=turn_array
            
            turn_array=turn_array+np.mean(self.CP_data_phone.acc_magnitude)
            

            self.b.plot(turn_array, color='red',alpha=0.5)
            # self.b.plot(turn_array/5, color='red',alpha=0.5)
            # if self.rightfoot or self.leftfoot:
            #     self.c.plot(turn_array/5, color='red',alpha=0.5)
            # if self.hand:
            #     self.d.plot(turn_array, color='red',alpha=0.5)

 
            self.canvas4.draw()


    def window_rms(self,a, window_size):
        a2 = np.power(a,2)
        window = np.ones(window_size)/float(window_size)
        return np.sqrt(np.convolve(a2, window, 'same'))
    
    
    def hl_envelopes_idx(self,s, dmin=1, dmax=1, split=False):
        """
        Input :
        s: 1d-array, data signal from which to extract high and low envelopes
        dmin, dmax: int, optional, size of chunks, use this if the size of the input signal is too big
        split: bool, optional, if True, split the signal in half along its mean, might help to generate the envelope in some cases
        Output :
        lmin,lmax : high/low envelope idx of input signal s
        """
    
        # locals min      
        lmin = (np.diff(np.sign(np.diff(s))) > 0).nonzero()[0] + 1 
        # locals max
        lmax = (np.diff(np.sign(np.diff(s))) < 0).nonzero()[0] + 1 
        
    
        if split:
            # s_mid is zero if s centered around x-axis or more generally mean of signal
            s_mid = np.mean(s) 
            # pre-sorting of locals min based on relative position with respect to s_mid 
            lmin = lmin[s[lmin]<s_mid]
            # pre-sorting of local max based on relative position with respect to s_mid 
            lmax = lmax[s[lmax]>s_mid]
    
    
        # global max of dmax-chunks of locals max 
        lmin = lmin[[i+np.argmin(s[lmin[i:i+dmin]]) for i in range(0,len(lmin),dmin)]]
        # global min of dmin-chunks of locals min 
        lmax = lmax[[i+np.argmax(s[lmax[i:i+dmax]]) for i in range(0,len(lmax),dmax)]]
        
        return lmin,lmax
            
            
    def ranges(self,nums,d=200):
        nums = sorted(set(nums))
        gaps = [[s, e] for s, e in zip(nums, nums[1:]) if s+d < e]
        edges = iter(nums[:1] + sum(gaps, []) + nums[-1:])
        return list(zip(edges, edges))
        
    def save_peaks(self,event=None):
        print("saving the peaks")
        np.savetxt(self.path+"\\preprocessed_acc_waist.csv",self.acc_w_magnitude,delimiter=",")
        np.savetxt(self.path+"\\preprocessed_gyro_waist.csv",self.gyro_w_magnitude,delimiter=",")
        
        if self.hand:
            np.savetxt(self.path+"\\preprocessed_acc_hand.csv",self.acc_h_magnitude,delimiter=",")
            np.savetxt(self.path+"\\preprocessed_gyro_hand.csv",self.gyro_h_magnitude,delimiter=",")

        
        np.savetxt(self.path+"\\steps_turn.csv",self.allsteps,delimiter=",")
        
        
    def run(self):
        self.mainwindow.mainloop()
   

if __name__ == '__main__':
    app = Label_data_app()
    app.run()